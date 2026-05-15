"""One-off: create regularisation Payrolls from two partner Excel files.

Usage:
    python manage.py create_regularisation_payrolls \\
        --bancobu-file <path> --ibb-file <path> \\
        --payment-plan-code "1.2 regular payment" \\
        [--dry-run]

Builds one Payroll per file, attached to a new PaymentCycle, with one
BenefitConsumption per resolved beneficiary at the exact amount from the
Excel. Routes through the Mera verification → approval → payment workflow.
"""
from __future__ import annotations

import csv
import re
import uuid
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl
from django.contrib.auth import get_user_model
from django.contrib.contenttypes.models import ContentType
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from contribution_plan.models import PaymentPlan
from individual.models import Group, GroupIndividual, Individual
from invoice.models import Bill
from payment_cycle.models import PaymentCycle
from payroll.models import (
    BenefitConsumption,
    BenefitConsumptionStatus,
    Payroll,
    PayrollStatus,
)
from payroll.services import PayrollService

from merankabandi.calcrule_strategies import MeraGroupBenefitPackageStrategy
from merankabandi.models import PaymentAgency

User = get_user_model()

FILES = (
    # (label, agency_code, candidate_sheet_names)
    # NB: the IBB Excel column says "IBB" but those payments route through
    # the INTERBANK PaymentAgency (BANCOBU has its own agency, IBB does not).
    ("bancobu", "BANCOBU",   ["Synthèse"]),
    ("ibb",     "INTERBANK", ["Non régularisés"]),
)


def _amount(raw):
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return int(raw)
    s = re.sub(r"[,\s\xa0]", "", str(raw).strip())
    try:
        return int(float(s)) if s else None
    except (ValueError, TypeError):
        return None


def _sid(raw):
    if raw is None:
        return None
    s = str(raw).strip().strip("\xa0").strip()
    return s or None


def _parse_excel(path, sheet_candidates):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = next((wb[s] for s in sheet_candidates if s in wb.sheetnames), wb.active)
    header = [
        str(h).strip().lower() if h else ""
        for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]

    def find(*candidates):
        for c in candidates:
            for i, h in enumerate(header):
                if c.lower() in h:
                    return i
        return None

    idx = {
        "sid":     find("social id"),
        "phone":   find("tel", "téléphone", "phone"),
        "nom":     find("noms", "nom"),
        "prenom":  find("prenoms", "prénom", "prenom"),
        "colline": find("colline"),
        "commune": find("commune"),
        "province": find("province"),
        "cni":     find("cni"),
        "amount":  find("montant"),
    }
    if idx["sid"] is None or idx["amount"] is None:
        raise CommandError(f"Couldn't find Social ID and Montant columns in {path}; header={header}")

    rows = []
    for r_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        sid = _sid(row[idx["sid"]])
        if not sid:
            continue

        def cell(key):
            j = idx[key]
            v = row[j] if j is not None else None
            return str(v).strip() if v is not None else ""

        rows.append({
            "excel_row": r_num,
            "social_id": sid,
            "nom":       cell("nom"),
            "prenom":    cell("prenom"),
            "cni":       cell("cni"),
            "phone":     cell("phone"),
            "province":  cell("province"),
            "commune":   cell("commune"),
            "colline":   cell("colline"),
            "amount":    _amount(row[idx["amount"]]),
        })
    return rows


def _dedupe(rows):
    """Return (kept, dropped_records). dropped_records = list of (row, outcome, note)."""
    by_sid = defaultdict(list)
    for i, r in enumerate(rows):
        by_sid[r["social_id"]].append(i)
    drop = {}
    for sid, positions in by_sid.items():
        if len(positions) <= 1:
            continue
        amounts = {rows[p]["amount"] for p in positions}
        if len(amounts) == 1:
            first_row = rows[positions[0]]["excel_row"]
            for p in positions[1:]:
                drop[p] = ("dedup_within_file", f"duplicate of row {first_row}")
        else:
            row_nums = [rows[p]["excel_row"] for p in positions]
            for p in positions:
                drop[p] = ("dedup_ambiguous", f"different amounts across rows {row_nums}")
    kept = [r for i, r in enumerate(rows) if i not in drop]
    dropped = [(rows[i], drop[i][0], drop[i][1]) for i in sorted(drop)]
    return kept, dropped


def _resolve_group(social_id):
    return (
        Group.objects.filter(code=social_id, is_deleted=False).first()
        or Group.objects.filter(
            json_ext__contains={"social_id": social_id}, is_deleted=False
        ).first()
    )


def _resolve_recipient(group):
    primary = GroupIndividual.objects.filter(
        group=group,
        recipient_type=GroupIndividual.RecipientType.PRIMARY,
        is_deleted=False,
    ).first()
    if primary:
        return primary.individual
    head = GroupIndividual.objects.filter(
        group=group, role=GroupIndividual.Role.HEAD, is_deleted=False,
    ).first()
    return head.individual if head else None


class Command(BaseCommand):
    help = "One-off: create regularisation Payrolls from two partner Excel files."

    def add_arguments(self, parser):
        # Either file is optional individually, but at least one must be given.
        parser.add_argument("--bancobu-file", default=None)
        parser.add_argument("--ibb-file", default=None)
        parser.add_argument("--payment-plan-code", required=True)
        parser.add_argument("--payment-cycle-code", default="1.2 — Régularisation 2026-05")
        parser.add_argument("--cycle-start-date", default="2026-05-11")
        parser.add_argument("--cycle-end-date", default="2026-05-31")
        parser.add_argument("--report-dir", default=".")
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--user", default=None,
                            help="Username; defaults to first superuser.")

    def handle(self, *args, **opts):
        user = self._get_user(opts.get("user"))

        try:
            payment_plan = PaymentPlan.objects.get(code=opts["payment_plan_code"])
        except PaymentPlan.DoesNotExist:
            available = list(
                PaymentPlan.objects.filter(is_deleted=False).values_list("code", flat=True)
            )
            raise CommandError(
                f"PaymentPlan code={opts['payment_plan_code']!r} not found.\n"
                f"Available codes:\n  " + "\n  ".join(available)
            )

        agencies = {}
        for _, code, _ in FILES:
            agency = PaymentAgency.objects.filter(
                code__iexact=code, is_active=True
            ).first()
            if not agency:
                raise CommandError(f"Active PaymentAgency {code!r} not found")
            agencies[code] = agency

        cycle_start = datetime.strptime(opts["cycle_start_date"], "%Y-%m-%d").date()
        cycle_end = datetime.strptime(opts["cycle_end_date"], "%Y-%m-%d").date()
        today = date.today()

        file_paths = {}
        if opts["bancobu_file"]:
            file_paths["bancobu"] = Path(opts["bancobu_file"])
        if opts["ibb_file"]:
            file_paths["ibb"] = Path(opts["ibb_file"])
        if not file_paths:
            raise CommandError("Provide at least one of --bancobu-file or --ibb-file")
        for p in file_paths.values():
            if not p.is_file():
                raise CommandError(f"File not found: {p}")

        sections = self._build_sections(file_paths, cycle_start, cycle_end, today)

        self._print_summary(sections)
        report_path = self._write_report(sections, Path(opts["report_dir"]))
        self.stdout.write(f"\nReport: {report_path}")

        if opts["dry_run"]:
            self.stdout.write(self.style.WARNING("Dry-run — nothing persisted."))
            return

        created = self._persist(sections, payment_plan, opts["payment_cycle_code"],
                                cycle_start, cycle_end, user, agencies)

        # Re-write report so it captures the DB-trigger-assigned BEN codes
        report_path = self._write_report(sections, Path(opts["report_dir"]))
        self.stdout.write(self.style.SUCCESS(f"\nReport: {report_path}"))
        self.stdout.write("\nPayrolls created:")
        for p in created:
            self.stdout.write(
                f"  {p.name}  uuid={p.id}  agency={p.json_ext['agency_code']}  status={p.status}"
            )

    def _get_user(self, username):
        # openIMIS custom User has no `is_superuser`; fall back to login_name='Admin'.
        if username:
            u = User.objects.filter(username=username).first()
            if not u:
                raise CommandError(f"User {username!r} not found")
            return u
        u = User.objects.filter(username__iexact="Admin").first()
        if not u:
            raise CommandError(
                "Pass --user <username> explicitly (no 'Admin' user found)."
            )
        return u

    def _build_sections(self, file_paths, cycle_start, cycle_end, today):
        sections = {}
        group_ct = ContentType.objects.get_for_model(Group)
        indiv_ct = ContentType.objects.get_for_model(Individual)

        for label, agency_code, sheets in FILES:
            if label not in file_paths:
                continue
            path = file_paths[label]
            rows = _parse_excel(path, sheets)
            kept, dropped = _dedupe(rows)
            section = {
                "label": label, "agency_code": agency_code, "path": path,
                "rows_total": len(rows), "entries": [], "triples": [],
            }
            for orig_row, outcome, note in dropped:
                section["entries"].append({
                    **orig_row, "file": label, "outcome": outcome, "note": note,
                    "group_id": "", "individual_id": "", "benefit_code": "",
                })

            for row in kept:
                entry = {**row, "file": label, "outcome": "", "note": "",
                         "group_id": "", "individual_id": "", "benefit_code": ""}

                if row["amount"] in (None, 0):
                    entry["outcome"] = "invalid_amount"
                    entry["note"] = "amount missing or zero"
                    section["entries"].append(entry)
                    continue

                group = _resolve_group(row["social_id"])
                if group is None:
                    entry["outcome"] = "unmatched_group"
                    entry["note"] = "social_id not in Group.code or Group.json_ext"
                    section["entries"].append(entry)
                    continue

                recipient = _resolve_recipient(group)
                if recipient is None:
                    entry["outcome"] = "no_recipient"
                    entry["note"] = "Group has no PRIMARY recipient and no HEAD"
                    entry["group_id"] = str(group.id)
                    section["entries"].append(entry)
                    continue

                entry["outcome"] = "created"
                entry["group_id"] = str(group.id)
                entry["individual_id"] = str(recipient.id)
                section["entries"].append(entry)

                bill_data = {
                    "subject_type_id":    group_ct.id, "subject_id":   group.id,
                    "thirdparty_type_id": indiv_ct.id, "thirdparty_id": recipient.id,
                    "code": "",
                    "date_due": cycle_end, "date_bill": today,
                    "date_valid_from": cycle_start, "date_valid_to": cycle_end,
                    "currency_tp_code": "BIF", "currency_code": "BIF",
                    "status": Bill.Status.VALIDATED,
                    "terms": "Régularisation 2026-05",
                }
                line_items = [{
                    "line_type_id": group_ct.id, "line_id": group.id,
                    "code": group.code, "quantity": 1,
                    "unit_price": row["amount"], "amount_total": row["amount"],
                    "date_valid_from": cycle_start, "date_valid_to": cycle_end,
                }]
                benefit_data = {
                    "individual_id": recipient.id, "code": "",
                    "date_due": cycle_end, "amount": row["amount"],
                    "type": "CASH", "status": BenefitConsumptionStatus.ACCEPTED,
                    "json_ext": {
                        "social_id":      row["social_id"],
                        "regularisation": True,
                        "source_file":    str(path),
                        "source_row":     row["excel_row"],
                        "agency_code":    agency_code,
                        "province":       row["province"],
                        "commune":        row["commune"],
                        "colline":        row["colline"],
                        "cni":            row["cni"],
                        "phone":          row["phone"],
                        "nom":            row["nom"],
                        "prenom":         row["prenom"],
                    },
                    "date_valid_from": cycle_start, "date_valid_to": cycle_end,
                }
                section["triples"].append(((bill_data, line_items, benefit_data), entry))

            sections[label] = section
        return sections

    def _persist(self, sections, payment_plan, cycle_code, cycle_start, cycle_end, user, agencies):
        created = []
        with transaction.atomic():
            cycle, _ = PaymentCycle.objects.get_or_create(
                code=cycle_code,
                defaults={
                    "start_date": cycle_start, "end_date": cycle_end,
                    "status": PaymentCycle.PaymentCycleStatus.ACTIVE,
                },
            )

            for label, agency_code, _ in FILES:
                if label not in sections:
                    continue
                section = sections[label]
                if not section["triples"]:
                    continue

                total = sum(t[0][2]["amount"] for t in section["triples"])
                # PaymentAgency.payment_gateway holds the strategy class name
                # (e.g. StrategyOnlinePaymentPull for BANCOBU,
                # StrategyOnlinePaymentPush for INTERBANK).
                payment_method = agencies[agency_code].payment_gateway
                payroll = Payroll(
                    id=uuid.uuid4(),
                    name=f"Régularisation {label.title()} 2026-05",
                    payment_plan=payment_plan,
                    payment_cycle=cycle,
                    payment_point=None,
                    status=PayrollStatus.GENERATING,
                    payment_method=payment_method,
                    json_ext={
                        "agency_code":    agency_code,
                        "regularisation": True,
                        "source_file":    str(section["path"]),
                        "row_count":      len(section["triples"]),
                        "total_amount":   str(total),
                    },
                )
                payroll.save(username=user.login_name)

                MeraGroupBenefitPackageStrategy.create_and_save_business_entities_batch(
                    [{"bill_data": b, "bill_data_line": l}
                     for ((b, l, _), _) in section["triples"]],
                    [{"benefit_data": bn} for ((_, _, bn), _) in section["triples"]],
                    payroll.id, user,
                )

                payroll.status = PayrollStatus.PENDING_APPROVAL
                payroll.save(username=user.login_name)
                PayrollService(user).create_accept_payroll_task(
                    payroll.id, {"id": str(payroll.id)},
                )
                # Signal `intercept_payroll_for_verification` now downgrades
                # status to PENDING_VERIFICATION and re-tags the task.
                payroll.refresh_from_db()

                # Backfill BEN-YY-XXXXXXXXXX codes into the entries
                bc_codes = list(BenefitConsumption.objects.filter(
                    payrollbenefitconsumption__payroll=payroll
                ).values_list("individual_id", "code"))
                code_by_indiv = {str(k): v for k, v in bc_codes}
                for ((_, _, _), entry) in section["triples"]:
                    entry["benefit_code"] = code_by_indiv.get(entry["individual_id"], "")

                created.append(payroll)
        return created

    def _print_summary(self, sections):
        self.stdout.write("=" * 80)
        self.stdout.write("Régularisation — création de payrolls")
        self.stdout.write("=" * 80)
        for label, _, _ in FILES:
            if label not in sections:
                continue
            s = sections[label]
            counts = Counter(e["outcome"] for e in s["entries"])
            total = sum((e["amount"] or 0) for e in s["entries"] if e["outcome"] == "created")
            self.stdout.write(
                f"{label:<8}  rows={s['rows_total']:>3}  "
                f"created={counts.get('created', 0):>3}  "
                f"dedup={counts.get('dedup_within_file', 0):>2}  "
                f"ambig={counts.get('dedup_ambiguous', 0):>2}  "
                f"unmatched={counts.get('unmatched_group', 0):>2}  "
                f"no_recip={counts.get('no_recipient', 0):>2}  "
                f"invalid_amt={counts.get('invalid_amount', 0):>2}  "
                f"total={total:>14,} BIF"
            )

    def _write_report(self, sections, report_dir):
        report_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d-%H%M%S")
        path = report_dir / f"regu_report_{ts}.csv"
        fields = [
            "file", "excel_row", "social_id", "nom", "prenom", "cni", "phone",
            "province", "commune", "colline", "amount", "outcome", "note",
            "group_id", "individual_id", "benefit_code",
        ]
        with path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
            writer.writeheader()
            for label, _, _ in FILES:
                if label not in sections:
                    continue
                for entry in sections[label]["entries"]:
                    writer.writerow({k: entry.get(k, "") for k in fields})
        return path
