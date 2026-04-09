"""
Import tracking/follow-up data from the Fiche de Suivi Excel file.

Processes ALL tracking sheets (107-col sheets with _uuid + tracking columns)
and the CollecteDesPlaintes sheet (resolution details: who/how resolved).

Updates existing tickets with:
- Status (oui → RESOLVED, en cours* → IN_PROGRESS)
- json_ext.tracking: responsible, dates, observation, resolution details

Usage:
    python manage.py import_suivi_tracking /path/to/Fiche_de_Suivi-2025.xlsx
    python manage.py import_suivi_tracking /path/to/file.xlsx --dry-run
    python manage.py import_suivi_tracking /path/to/file.xlsx --sheet Suivi  # single sheet only
"""
import json
import logging
from datetime import datetime

import openpyxl
from django.core.management.base import BaseCommand
from django.db import connection

logger = logging.getLogger(__name__)

# Tracking sheets (107-col format with _uuid + Plainte résolue/Responsable/etc.)
TRACKING_SHEETS = ['Suivi', 'Feuil2', 'Plainte-trim2-2025', 'Plainte trim3', 'Plaintes trim 4', 'Feuil3']

# Sheet with resolution narrative (who resolved, how)
RESOLUTION_SHEET = 'CollecteDesPlaintes(version'


class Command(BaseCommand):
    help = 'Import tracking data from Fiche de Suivi Excel into existing tickets'

    def add_arguments(self, parser):
        parser.add_argument('file', help='Path to Fiche de Suivi Excel file')
        parser.add_argument('--dry-run', action='store_true', help='Preview changes without saving')
        parser.add_argument('--sheet', default=None,
                            help='Process a single sheet only (default: all tracking sheets + CollecteDesPlaintes)')

    def handle(self, *args, **options):
        file_path = options['file']
        dry_run = options['dry_run']
        single_sheet = options.get('sheet')

        self.stdout.write(f"Loading {file_path}...")
        wb = openpyxl.load_workbook(file_path, data_only=True)

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN — no changes will be applied"))

        # Load all DB tickets
        db_tickets = self._load_db_tickets()
        self.stdout.write(f"DB tickets: {len(db_tickets)}")

        # Determine which sheets to process
        if single_sheet:
            sheets_to_process = [single_sheet]
        else:
            sheets_to_process = [s for s in TRACKING_SHEETS if s in wb.sheetnames]

        # Phase 1: Process tracking sheets (best data wins per uuid)
        merged_tracking = {}
        for sheet_name in sheets_to_process:
            rows = self._read_tracking_sheet(wb[sheet_name])
            self.stdout.write(f"  {sheet_name}: {len(rows)} rows")
            for row in rows:
                uuid = row['uuid']
                existing = merged_tracking.get(uuid)
                if not existing:
                    merged_tracking[uuid] = row
                else:
                    # Merge: fill gaps from this sheet
                    for field in ['resolved', 'responsible', 'date_submitted', 'date_resolved',
                                  'resolution_delay_days', 'observation']:
                        if not existing.get(field) and row.get(field):
                            existing[field] = row[field]

        self.stdout.write(f"\nMerged tracking data: {len(merged_tracking)} unique tickets")

        # Phase 2: Read CollecteDesPlaintes resolution details
        resolution_data = {}
        if not single_sheet and RESOLUTION_SHEET in wb.sheetnames:
            resolution_data = self._read_resolution_sheet(wb[RESOLUTION_SHEET])
            self.stdout.write(f"CollecteDesPlaintes resolution data: {len(resolution_data)} tickets")

        # Phase 3: Apply updates
        updated = 0
        status_changed = 0
        resolution_enriched = 0
        not_found = 0
        skipped = 0

        # Process tracking data
        for uuid, row in merged_tracking.items():
            if uuid not in db_tickets:
                not_found += 1
                continue

            db = db_tickets[uuid]
            existing_ext = db['json_ext']
            existing_tracking = existing_ext.get('tracking', {})

            # Skip if already fully tracked (idempotent)
            if (existing_tracking.get('date_resolved') and existing_tracking.get('responsible')
                    and existing_tracking.get('resolver_name')):
                skipped += 1
                continue

            # Build/update tracking section
            tracking = {**existing_tracking}
            for field in ['responsible', 'date_submitted', 'date_resolved',
                          'resolution_delay_days', 'observation', 'resolved_raw']:
                if row.get(field) and not tracking.get(field):
                    tracking[field] = row[field]

            # Merge resolution details from CollecteDesPlaintes
            res = resolution_data.get(uuid, {})
            if res:
                if res.get('resolver') and not tracking.get('resolver_name'):
                    tracking['resolver_name'] = res['resolver']
                    resolution_enriched += 1
                if res.get('resolver_function') and not tracking.get('resolver_function'):
                    tracking['resolver_function'] = res['resolver_function']
                if res.get('how_resolved') and not tracking.get('how_resolved'):
                    tracking['how_resolved'] = res['how_resolved']

            new_ext = {**existing_ext, 'tracking': tracking}

            # Determine status
            new_status = db['status']
            resolved_val = row.get('resolved')
            if resolved_val in ('oui', 'résolu'):
                new_status = 'RESOLVED'
            elif resolved_val and 'en cours' in resolved_val:
                new_status = 'IN_PROGRESS'

            status_will_change = new_status != db['status']

            if not dry_run:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        UPDATE grievance_social_protection_ticket
                        SET "Json_ext" = %s, status = %s, "DateUpdated" = NOW()
                        WHERE "UUID" = %s
                    """, [json.dumps(new_ext), new_status, uuid])

            updated += 1
            if status_will_change:
                status_changed += 1

        # Phase 4: Apply resolution-only data (tickets in CollecteDesPlaintes but NOT in tracking sheets)
        resolution_only = 0
        for uuid, res in resolution_data.items():
            if uuid in merged_tracking:
                continue  # Already handled above
            if uuid not in db_tickets:
                continue
            if not res.get('resolver') and not res.get('how_resolved'):
                continue

            db = db_tickets[uuid]
            existing_ext = db['json_ext']
            existing_tracking = existing_ext.get('tracking', {})

            if existing_tracking.get('resolver_name') and existing_tracking.get('how_resolved'):
                continue  # Already has this data

            tracking = {**existing_tracking}
            if res.get('resolver') and not tracking.get('resolver_name'):
                tracking['resolver_name'] = res['resolver']
            if res.get('resolver_function') and not tracking.get('resolver_function'):
                tracking['resolver_function'] = res['resolver_function']
            if res.get('how_resolved') and not tracking.get('how_resolved'):
                tracking['how_resolved'] = res['how_resolved']

            new_ext = {**existing_ext, 'tracking': tracking}

            if not dry_run:
                with connection.cursor() as cursor:
                    cursor.execute("""
                        UPDATE grievance_social_protection_ticket
                        SET "Json_ext" = %s, "DateUpdated" = NOW()
                        WHERE "UUID" = %s
                    """, [json.dumps(new_ext), uuid])

            resolution_only += 1

        self.stdout.write(self.style.SUCCESS(
            f"\n{'[DRY RUN] ' if dry_run else ''}"
            f"Tracking updated: {updated}, Status changed: {status_changed}, "
            f"Resolution enriched: {resolution_enriched}, Resolution-only: {resolution_only}, "
            f"Skipped (already complete): {skipped}, Not found: {not_found}"
        ))

    def _load_db_tickets(self):
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT "UUID"::text, status, "Json_ext"
                FROM grievance_social_protection_ticket
                WHERE "isDeleted" = false
            """)
            tickets = {}
            for row in cursor.fetchall():
                ext = row[2]
                if isinstance(ext, str):
                    try:
                        ext = json.loads(ext)
                    except (json.JSONDecodeError, TypeError):
                        ext = {}
                tickets[row[0]] = {'status': row[1], 'json_ext': ext or {}}
            return tickets

    def _read_tracking_sheet(self, ws):
        """Read a 107-col tracking sheet. Returns list of row dicts."""
        col_map = {}
        for c in ws[1]:
            if c.value:
                col_map[str(c.value).strip()] = c.column

        uuid_col = col_map.get('_uuid')
        if not uuid_col:
            return []

        resolved_col = next((col_map[h] for h in col_map if 'plainte résolue' in h.lower()), None)
        resp_col = next((col_map[h] for h in col_map if h.lower() == 'responsable'), None)
        date_sub_col = next((col_map[h] for h in col_map if 'date de soumission' in h.lower()), None)
        date_res_col = next((col_map[h] for h in col_map if 'date de résolution' in h.lower()), None)
        delay_col = next((col_map[h] for h in col_map if 'délai de résolution' in h.lower()), None)
        obs_col = next((col_map[h] for h in col_map if h.lower() == 'observation'), None)

        rows = []
        for row_idx in range(2, ws.max_row + 1):
            uuid_val = ws.cell(row=row_idx, column=uuid_col).value
            if not uuid_val:
                continue

            resolved = ws.cell(row=row_idx, column=resolved_col).value if resolved_col else None
            responsible = ws.cell(row=row_idx, column=resp_col).value if resp_col else None
            date_sub = ws.cell(row=row_idx, column=date_sub_col).value if date_sub_col else None
            date_res = ws.cell(row=row_idx, column=date_res_col).value if date_res_col else None
            delay = ws.cell(row=row_idx, column=delay_col).value if delay_col else None
            observation = ws.cell(row=row_idx, column=obs_col).value if obs_col else None

            rows.append({
                'uuid': str(uuid_val).strip(),
                'resolved': str(resolved).strip().lower() if resolved else None,
                'responsible': str(responsible).strip() if responsible else None,
                'date_submitted': _parse_date(date_sub),
                'date_resolved': _parse_date(date_res),
                'resolution_delay_days': _parse_int(delay),
                'observation': str(observation).strip() if observation else None,
            })

        return rows

    def _read_resolution_sheet(self, ws):
        """Read CollecteDesPlaintes resolution columns (93-96). Returns dict by uuid."""
        col_map = {}
        for c in ws[1]:
            if c.value:
                col_map[str(c.value).strip()[:60]] = c.column

        uuid_col = col_map.get('_uuid')
        if not uuid_col:
            return {}

        resolved_col = next((col_map[h] for h in col_map if 'plainte a déjà été résolue' in h.lower()), None)
        resolver_col = next((col_map[h] for h in col_map if 'qui a résolu' in h.lower()), None)
        func_col = next((col_map[h] for h in col_map if 'quelle est sa fonction' in h.lower()), None)
        how_col = next((col_map[h] for h in col_map if 'comment la plainte' in h.lower()), None)

        data = {}
        for row_idx in range(2, ws.max_row + 1):
            uuid_val = ws.cell(row=row_idx, column=uuid_col).value
            if not uuid_val:
                continue

            resolver = ws.cell(row=row_idx, column=resolver_col).value if resolver_col else None
            func = ws.cell(row=row_idx, column=func_col).value if func_col else None
            how = ws.cell(row=row_idx, column=how_col).value if how_col else None

            entry = {}
            if resolver and str(resolver).strip().lower() not in ('none', ''):
                entry['resolver'] = str(resolver).strip()
            if func and str(func).strip().lower() not in ('none', ''):
                entry['resolver_function'] = str(func).strip()
            if how and str(how).strip().lower() not in ('none', ''):
                entry['how_resolved'] = str(how).strip()

            if entry:
                data[str(uuid_val).strip()] = entry

        return data


def _parse_date(val):
    """Parse date from Excel — could be datetime, string, or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s.lower() == 'none':
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s


def _parse_int(val):
    """Parse integer from Excel — could be int, float, string, or None."""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None
