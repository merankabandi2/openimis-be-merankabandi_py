"""Merankabandi-specific Excel export for group beneficiaries.

Registered via ``social_protection.export_mixin.register_export_handler`` from
``MerankabandiConfig.ready()`` so that the upstream ``ExportableSocialProtectionQueryMixin``
dispatches XLSX exports of ``group_beneficiary`` to this module without any
merankabandi-specific code living in the upstream package.

Custom shape over the stock CSV export:
  - Single row per beneficiary household (only the PRIMARY recipient)
  - Province / Commune / Colline resolved from the location hierarchy
  - Demographic fields pulled from ``individual.json_ext`` (sexe, ci, pere,
    mere, telephone) and ``group.json_ext.moyen_telecom.msisdn``
  - Per-row photo URLs (recipient photo + CNI recto/verso) pointing at
    ``/api/merankabandi/beneficiary-photo/<type>/<individual_uuid>/``
"""
from __future__ import annotations

import datetime
import logging
from io import BytesIO

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage

from core.models import ExportableQueryModel

logger = logging.getLogger(__name__)


class BeneficiaryExcelExportService:
    """Build an Excel workbook of group beneficiaries with photo URLs.

    The ``base_url`` default is the production hostname so a CLI invocation
    (e.g. management command) still produces clickable links; the registered
    GQL handler overrides it with ``info.context.build_absolute_uri('/')``.
    """

    def __init__(self, user, base_url: str = 'https://mis.merankabandi2.bi'):
        self.user = user
        self.base_url = base_url.rstrip('/')

    def export_group_beneficiaries_to_excel(self, group_beneficiaries_queryset) -> BytesIO:
        from openpyxl import Workbook
        from individual.models import GroupIndividual

        wb = Workbook()
        ws = wb.active
        ws.title = "Beneficiaries"

        headers = [
            'nom', 'prenom', 'province', 'commune', 'colline', 'cni',
            'naissance_date', 'genre', 'socialid', 'pere', 'mere', 'phone',
            'photo', 'cni_recto', 'cni_verso', 'personnal_phone', 'mutwa',
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        row_num = 2
        for gb in group_beneficiaries_queryset.select_related(
            'group',
            'group__location',
            'group__location__parent',
            'group__location__parent__parent',
        ).prefetch_related('group__groupindividuals__individual'):

            colline = gb.group.location if gb.group.location else None
            commune = colline.parent if colline and colline.parent else None
            province = commune.parent if commune and commune.parent else None

            primary_recipient = GroupIndividual.objects.filter(
                group=gb.group,
                recipient_type='PRIMARY',
                is_deleted=False,
            ).select_related('individual').first()
            if not primary_recipient:
                continue

            individual = primary_recipient.individual
            ij = individual.json_ext or {}
            gj = gb.json_ext or {}

            naissance_date = individual.dob.strftime('%Y-%m-%d') if individual.dob else ''
            socialid = gb.group.code if gb.group else ''
            phone = (gj.get('moyen_telecom') or {}).get('msisdn', '') if 'moyen_telecom' in gj else ''

            uuid_str = str(individual.id)
            photo_url = f"{self.base_url}/api/merankabandi/beneficiary-photo/photo/{uuid_str}/"
            cni_recto_url = f"{self.base_url}/api/merankabandi/beneficiary-photo/photo_ci1/{uuid_str}/"
            cni_verso_url = f"{self.base_url}/api/merankabandi/beneficiary-photo/photo_ci2/{uuid_str}/"

            row_data = [
                individual.last_name,
                individual.first_name,
                province.name if province else '',
                commune.name if commune else '',
                colline.name if colline else '',
                ij.get('ci', ''),
                naissance_date,
                ij.get('sexe', ''),
                socialid,
                ij.get('pere', ''),
                ij.get('mere', ''),
                phone,
                photo_url,
                cni_recto_url,
                cni_verso_url,
                ij.get('telephone', ''),
                gj.get('menage_mutwa', ''),
            ]
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
            row_num += 1

        # Auto-adjust column widths (capped at 50 chars)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        return out


def export_group_beneficiary_xlsx(queryset, user, info) -> ExportableQueryModel:
    """Registry handler — `social_protection.export_mixin.register_export_handler`-compatible.

    Builds the XLSX via ``BeneficiaryExcelExportService``, saves it through
    Django's default storage, records an ``ExportableQueryModel`` and returns
    that model. The mixin reads ``.name`` for the GQL response.
    """
    base_url = info.context.build_absolute_uri('/').rstrip('/')
    excel_buf = BeneficiaryExcelExportService(user, base_url) \
        .export_group_beneficiaries_to_excel(queryset)

    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'beneficiaries_export_{timestamp}.xlsx'
    saved_path = default_storage.save(
        f'exports/{filename}', ContentFile(excel_buf.getvalue()),
    )

    export_obj = ExportableQueryModel(
        name=filename,
        model='GroupBeneficiary',
        content=saved_path,
        user=user,
        sql_query=str(queryset.query),
        file_format='xlsx',
    )
    export_obj.save()
    return export_obj


# ─── REST endpoint (admin / direct-download path) ─────────────────────────────
# Streams the XLSX directly in the HTTP response without creating an
# ExportableQueryModel row. Previously lived in
# ``social_protection/views.py::export_beneficiaries_excel``; moved here so
# the upstream package stays merankabandi-free.

from rest_framework.decorators import api_view, permission_classes
from core.views import check_user_rights
from social_protection.apps import SocialProtectionConfig


@api_view(["GET"])
@permission_classes(
    [check_user_rights(SocialProtectionConfig.gql_beneficiary_search_perms, )]
)
def export_beneficiaries_excel_view(request):
    """Direct-download XLSX export of GroupBeneficiary filtered by benefit_plan_id + status."""
    from django.http import HttpResponse
    from rest_framework import status as drf_status
    from rest_framework.response import Response
    from social_protection.models import GroupBeneficiary

    try:
        benefit_plan_id = request.query_params.get('benefit_plan_id')
        status_filter = request.query_params.get('status')

        queryset = GroupBeneficiary.objects.filter(is_deleted=False)
        if benefit_plan_id:
            queryset = queryset.filter(benefit_plan_id=benefit_plan_id)
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        queryset = queryset.order_by('-date_created', 'id')

        base_url = request.build_absolute_uri('/').rstrip('/')
        excel_buf = BeneficiaryExcelExportService(request.user, base_url) \
            .export_group_beneficiaries_to_excel(queryset)

        response = HttpResponse(
            excel_buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="beneficiaries_{status_filter or "all"}.xlsx"'
        )
        return response
    except Exception as exc:
        logger.error("Error while exporting group beneficiaries to Excel", exc_info=exc)
        return Response(
            {'success': False, 'error': str(exc)},
            status=drf_status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
