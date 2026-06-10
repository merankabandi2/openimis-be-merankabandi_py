from datetime import date
from io import BytesIO

import openpyxl
from django.test import TestCase
from rest_framework.test import APIRequestFactory, force_authenticate

from core.test_helpers import create_test_interactive_user
from social_protection.models import BenefitPlan, GroupBeneficiary
from individual.models import Group, GroupIndividual, Individual
from location.models import Location

from merankabandi.models import PaymentAgency, ProvincePaymentAgency
from merankabandi.reports.account_creation_report import (
    AccountCreationReportService, normalize_genre, sanitize_sheet_title, REPORT_HEADERS,
)


class TestHelpers(TestCase):
    def test_normalize_genre(self):
        self.assertEqual(normalize_genre('F'), 'F')
        self.assertEqual(normalize_genre('Féminin'), 'F')
        self.assertEqual(normalize_genre('Male'), 'M')
        self.assertEqual(normalize_genre(''), '')
        self.assertEqual(normalize_genre(None), '')

    def test_sanitize_sheet_title_truncates_and_strips(self):
        used = set()
        t = sanitize_sheet_title('A/Very:Long*Commune?Name[That]Exceeds31Chars', used)
        self.assertLessEqual(len(t), 31)
        for ch in '[]:*?/\\':
            self.assertNotIn(ch, t)

    def test_sanitize_sheet_title_dedupes(self):
        used = set()
        a = sanitize_sheet_title('Kinyinya', used)
        b = sanitize_sheet_title('Kinyinya', used)
        self.assertNotEqual(a.lower(), b.lower())


def _user():
    return create_test_interactive_user(username='reportuser')


def _bp(user):
    bp = BenefitPlan(code='RPT-12', name='Report Plan',
                     type=BenefitPlan.BenefitPlanType.GROUP_TYPE, date_valid_from=date(2024, 1, 1))
    bp.save(username=user.login_name)
    return bp


def _loc(name, type_, parent=None):
    loc = Location(name=name, code=name[:8], type=type_, parent=parent)
    loc.save()
    return loc


def _beneficiary(user, bp, colline, *, social_id, first, last, sexe='M', ci='C1',
                 moyen_paiement=None):
    g = Group(code=social_id, location=colline, json_ext={})
    g.save(username=user.login_name)
    ind = Individual(first_name=first, last_name=last, dob='1980-01-01',
                     json_ext={'sexe': sexe, 'ci': ci})
    ind.save(username=user.login_name)
    gi = GroupIndividual(group=g, individual=ind,
                         role=GroupIndividual.Role.HEAD,
                         recipient_type=GroupIndividual.RecipientType.PRIMARY)
    gi.save(username=user.login_name)
    je = {}
    if moyen_paiement is not None:
        je['moyen_paiement'] = moyen_paiement
    gb = GroupBeneficiary(group=g, benefit_plan=bp, status='ACTIVE', json_ext=je)
    gb.save(username=user.login_name)
    return gb


class TestAgencyScoping(TestCase):
    """S2 (cross-review): attribution endpoints must scope to the agency's served
    provinces/plans, not return program-wide beneficiary PII."""

    def setUp(self):
        self.user = _user()
        self.bp = _bp(self.user)
        self.prov_a = _loc('ProvA', 'D')
        self.prov_b = _loc('ProvB', 'D')
        com_a = _loc('ComA', 'W', self.prov_a)
        com_b = _loc('ComB', 'W', self.prov_b)
        self.col_a = _loc('ColA', 'V', com_a)
        self.col_b = _loc('ColB', 'V', com_b)
        _beneficiary(self.user, self.bp, self.col_a, social_id='A1', first='A', last='A')
        _beneficiary(self.user, self.bp, self.col_b, social_id='B1', first='B', last='B')
        self.agency = PaymentAgency.objects.create(
            code='AGX', name='AGX', payment_gateway='StrategyOnlinePaymentPull', is_active=True)
        # Agency AGX serves only province A for this plan.
        ProvincePaymentAgency.objects.create(
            province=self.prov_a, benefit_plan=self.bp, payment_agency=self.agency)

    def test_scopes_to_served_province_only(self):
        from social_protection.models import GroupBeneficiary
        from merankabandi.services import _scope_beneficiaries_to_agency
        qs = _scope_beneficiaries_to_agency(GroupBeneficiary.objects.all(), 'AGX')
        codes = sorted(gb.group.code for gb in qs)
        self.assertEqual(codes, ['A1'])  # NOT B1 (province B not served by AGX)

    def test_unknown_agency_returns_nothing(self):
        from social_protection.models import GroupBeneficiary
        from merankabandi.services import _scope_beneficiaries_to_agency
        qs = _scope_beneficiaries_to_agency(GroupBeneficiary.objects.all(), 'NOPE')
        self.assertEqual(list(qs), [])

    def test_no_agency_passes_through(self):
        from social_protection.models import GroupBeneficiary
        from merankabandi.services import _scope_beneficiaries_to_agency
        qs = _scope_beneficiaries_to_agency(GroupBeneficiary.objects.all(), None)
        self.assertEqual(qs.count(), GroupBeneficiary.objects.count())


class TestAccountCreationReport(TestCase):
    def setUp(self):
        self.user = _user()
        self.bp = _bp(self.user)
        self.province = _loc('Gitega', 'D')
        self.commune_a = _loc('Giheta', 'W', self.province)
        self.commune_b = _loc('Gishubi', 'W', self.province)
        self.colline_a = _loc('CollineA', 'V', self.commune_a)
        self.colline_b = _loc('CollineB', 'V', self.commune_b)
        _beneficiary(self.user, self.bp, self.colline_a, social_id='S1', first='Jean', last='UWIMANA',
                     sexe='M', ci='111',
                     moyen_paiement={'phoneNumber': '71519246', 'agence': 'ECONET',
                                     'status': 'SUCCESS', 'ordernumber': 2852,
                                     'responseDate': '2023-06-23'})
        _beneficiary(self.user, self.bp, self.colline_a, social_id='S2', first='Marie', last='NDAYI',
                     sexe='F', ci='222', moyen_paiement=None)
        _beneficiary(self.user, self.bp, self.colline_b, social_id='S3', first='Paul', last='BIZIMANA',
                     sexe='M', ci='333',
                     moyen_paiement={'phoneNumber': '79045797', 'agence': 'LUMITEL',
                                     'status': 'SUCCESS', 'ordernumber': 99,
                                     'responseDate': '2023-07-01'})

    def _load(self, buf):
        return openpyxl.load_workbook(BytesIO(buf.getvalue()))

    def test_resolve_communes_by_province(self):
        svc = AccountCreationReportService(self.user)
        communes = list(svc.resolve_communes(self.bp.id, province_id=self.province.id))
        names = sorted(c.name for c in communes)
        self.assertEqual(names, ['Giheta', 'Gishubi'])

    def test_resolve_communes_by_payment_agency(self):
        agency = PaymentAgency.objects.create(code='RPT-FINBANK', name='RPT-FINBANK',
                                              payment_gateway='StrategyOnlinePaymentPull', is_active=True)
        ProvincePaymentAgency.objects.create(province=self.province, benefit_plan=self.bp,
                                             payment_agency=agency)
        svc = AccountCreationReportService(self.user)
        communes = list(svc.resolve_communes(self.bp.id, payment_agency_id=agency.id))
        self.assertEqual(sorted(c.name for c in communes), ['Giheta', 'Gishubi'])

    def test_resolve_communes_excludes_inactive_agency_assignment(self):
        agency = PaymentAgency.objects.create(code='RPT-INACT', name='RPT-INACT',
                                              payment_gateway='StrategyOnlinePaymentPull', is_active=True)
        # A deactivated province<->agency assignment must NOT contribute communes.
        ProvincePaymentAgency.objects.create(province=self.province, benefit_plan=self.bp,
                                             payment_agency=agency, is_active=False)
        svc = AccountCreationReportService(self.user)
        communes = list(svc.resolve_communes(self.bp.id, payment_agency_id=agency.id))
        self.assertEqual(communes, [])

    def test_workbook_has_a_sheet_per_commune(self):
        svc = AccountCreationReportService(self.user)
        buf = svc.build_workbook(self.bp.id, province_id=self.province.id)
        wb = self._load(buf)
        self.assertEqual(set(wb.sheetnames), {'Giheta', 'Gishubi'})

    def test_headers_and_success_row(self):
        svc = AccountCreationReportService(self.user)
        wb = self._load(svc.build_workbook(self.bp.id, province_id=self.province.id))
        ws = wb['Giheta']
        self.assertEqual([c.value for c in ws[1]], REPORT_HEADERS)
        rows = {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True)}
        s1 = rows['S1']
        d = dict(zip(REPORT_HEADERS, s1))
        self.assertEqual(d['nom'], 'UWIMANA')
        self.assertEqual(d['prenom'], 'Jean')
        self.assertEqual(d['genre'], 'M')
        self.assertEqual(d['commune'], 'Giheta')
        self.assertEqual(d['province'], 'Gitega')
        self.assertEqual(d['telephone'], '71519246')
        self.assertEqual(d['operateur_mobile'], 'ECONET')
        self.assertEqual(d['statut_compte'], 'SUCCESS')
        self.assertEqual(str(d['numero_ordre']), '2852')

    def test_aucun_when_no_moyen_paiement(self):
        svc = AccountCreationReportService(self.user)
        wb = self._load(svc.build_workbook(self.bp.id, province_id=self.province.id))
        ws = wb['Giheta']
        rows = {r[0]: dict(zip(REPORT_HEADERS, r)) for r in ws.iter_rows(min_row=2, values_only=True)}
        self.assertEqual(rows['S2']['statut_compte'], 'AUCUN')
        self.assertIn(rows['S2']['telephone'], (None, ''))

    def test_empty_commune_has_header_only_sheet(self):
        empty_commune = _loc('Makebuko', 'W', self.province)
        svc = AccountCreationReportService(self.user)
        wb = self._load(svc.build_workbook(self.bp.id, province_id=self.province.id))
        self.assertIn('Makebuko', wb.sheetnames)
        ws = wb['Makebuko']
        self.assertEqual([c.value for c in ws[1]], REPORT_HEADERS)
        self.assertEqual(ws.max_row, 1)


class TestAccountCreationReportView(TestCase):
    def setUp(self):
        from merankabandi.views_reports import account_creation_report_view
        self.view = account_creation_report_view
        self.factory = APIRequestFactory()
        self.user = _user()
        self.bp = _bp(self.user)
        self.province = _loc('Gitega', 'D')
        self.commune = _loc('Giheta', 'W', self.province)
        self.colline = _loc('CollineA', 'V', self.commune)
        _beneficiary(self.user, self.bp, self.colline, social_id='S1', first='Jean', last='UWIMANA',
                     moyen_paiement={'phoneNumber': '71519246', 'agence': 'ECONET',
                                     'status': 'SUCCESS', 'ordernumber': 1, 'responseDate': '2023-01-01'})

    def _get(self, params):
        # Plain RequestFactory triggers SessionAuthentication CSRF enforcement;
        # use DRF's APIRequestFactory + force_authenticate so the permission layer
        # (IsAuthenticated + has_perms) resolves request.user to the test user.
        req = self.factory.get('/api/merankabandi/reports/account-creation/', params)
        force_authenticate(req, user=self.user)
        return self.view(req)

    def test_missing_benefit_plan_returns_400(self):
        resp = self._get({'province_id': str(self.province.id)})
        self.assertEqual(resp.status_code, 400)

    def test_both_scopes_returns_400(self):
        resp = self._get({'benefit_plan_id': str(self.bp.id),
                          'province_id': str(self.province.id),
                          'payment_agency_id': '1'})
        self.assertEqual(resp.status_code, 400)

    def test_valid_returns_xlsx(self):
        resp = self._get({'benefit_plan_id': str(self.bp.id),
                          'province_id': str(self.province.id)})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'],
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('attachment;', resp['Content-Disposition'])
