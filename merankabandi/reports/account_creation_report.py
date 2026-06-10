"""Per-commune Finbank account-creation report.

One worksheet per commune; one row per GroupBeneficiary of a program, with the
beneficiary's payment-account attribution status (moyen_paiement). Read-only.

Scope = program (benefit plan) + (province OR payment agency). Payment-agency
scope resolves communes via ProvincePaymentAgency (province+benefit_plan+agency).
"""
from __future__ import annotations

import logging
import re
from io import BytesIO

logger = logging.getLogger(__name__)

REPORT_HEADERS = [
    'socialid', 'nom', 'prenom', 'cni', 'naissance_date', 'genre',
    'province', 'commune', 'colline', 'telephone', 'operateur_mobile',
    'statut_compte', 'numero_ordre', 'date_attribution',
]

_ILLEGAL_SHEET_CHARS = re.compile(r'[\[\]:*?/\\]')


def normalize_genre(value) -> str:
    if not value:
        return ''
    v = str(value).strip().lower()
    if v in ('f', 'féminin', 'female'):
        return 'F'
    if v in ('m', 'masculin', 'male'):
        return 'M'
    return ''


def sanitize_sheet_title(name: str, used: set) -> str:
    """Excel sheet title: <=31 chars, no []:*?/\\, unique (case-insensitive)."""
    title = _ILLEGAL_SHEET_CHARS.sub(' ', name or 'Commune').strip()[:31] or 'Commune'
    base = title
    i = 1
    while title.lower() in used:
        suffix = f'~{i}'
        title = (base[:31 - len(suffix)] + suffix)
        i += 1
    used.add(title.lower())
    return title


class AccountCreationReportService:
    """Builds the multi-sheet account-creation workbook."""

    def __init__(self, user):
        self.user = user

    def resolve_communes(self, benefit_plan_id, province_id=None, payment_agency_id=None):
        """Return a queryset of commune Locations (type 'W') in scope, ordered."""
        from location.models import Location
        from merankabandi.models import ProvincePaymentAgency

        if province_id:
            qs = Location.objects.filter(
                type='W', parent_id=province_id, validity_to__isnull=True)
        elif payment_agency_id:
            province_ids = list(ProvincePaymentAgency.objects.filter(
                payment_agency_id=payment_agency_id,
                benefit_plan_id=benefit_plan_id,
                is_active=True,
            ).values_list('province_id', flat=True))
            qs = Location.objects.filter(
                type='W', parent_id__in=province_ids, validity_to__isnull=True)
        else:
            qs = Location.objects.none()
        return qs.order_by('parent__name', 'name')

    def estimate_size(self, benefit_plan_id, province_id=None, payment_agency_id=None):
        """Cheap beneficiary count for the scope — drives the sync/async decision
        without building the workbook (one COUNT query)."""
        from social_protection.models import GroupBeneficiary

        commune_ids = list(self.resolve_communes(
            benefit_plan_id, province_id, payment_agency_id).values_list('id', flat=True))
        if not commune_ids:
            return 0
        return GroupBeneficiary.objects.filter(
            benefit_plan_id=benefit_plan_id, is_deleted=False,
            group__location__parent_id__in=commune_ids,
        ).count()

    def _beneficiaries_for_commune(self, benefit_plan_id, commune):
        from social_protection.models import GroupBeneficiary
        return (GroupBeneficiary.objects
                .filter(benefit_plan_id=benefit_plan_id, is_deleted=False,
                        group__location__parent_id=commune.id)
                .select_related('group', 'group__location', 'group__location__parent',
                                'group__location__parent__parent')
                .order_by('group__code'))

    def _row(self, gb):
        from individual.models import GroupIndividual

        colline = gb.group.location if gb.group else None
        commune = colline.parent if colline and colline.parent else None
        province = commune.parent if commune and commune.parent else None

        primary = (GroupIndividual.objects
                   .filter(group=gb.group, recipient_type='PRIMARY', is_deleted=False)
                   .select_related('individual').first())
        if not primary:
            primary = (GroupIndividual.objects
                       .filter(group=gb.group, role=GroupIndividual.Role.HEAD, is_deleted=False)
                       .select_related('individual').first())
        ind = primary.individual if primary else None
        ij = (ind.json_ext or {}) if ind else {}
        gj = gb.json_ext or {}
        mp = gj.get('moyen_paiement') if isinstance(gj.get('moyen_paiement'), dict) else None
        mt = gj.get('moyen_telecom') if isinstance(gj.get('moyen_telecom'), dict) else None

        telephone = (mp or {}).get('phoneNumber') or (mt or {}).get('msisdn') or ''
        statut = (mp or {}).get('status') if mp else None
        statut = statut or 'AUCUN'
        return [
            gb.group.code if gb.group else '',
            ind.last_name if ind else '',
            ind.first_name if ind else '',
            ij.get('ci', ''),
            ind.dob.strftime('%Y-%m-%d') if ind and ind.dob else '',
            normalize_genre(ij.get('sexe')),
            province.name if province else '',
            commune.name if commune else '',
            colline.name if colline else '',
            telephone,
            (mp or {}).get('agence', '') if mp else '',
            statut,
            (mp or {}).get('ordernumber', '') if mp else '',
            (mp or {}).get('responseDate') or (mp or {}).get('requestDate') or '' if mp else '',
        ]

    def build_workbook(self, benefit_plan_id, province_id=None, payment_agency_id=None) -> BytesIO:
        from openpyxl import Workbook

        communes = self.resolve_communes(benefit_plan_id, province_id, payment_agency_id)
        wb = Workbook()
        wb.remove(wb.active)
        used_titles = set()
        total_rows = 0
        commune_count = 0
        for commune in communes:
            commune_count += 1
            ws = wb.create_sheet(sanitize_sheet_title(commune.name, used_titles))
            ws.append(REPORT_HEADERS)
            for gb in self._beneficiaries_for_commune(benefit_plan_id, commune):
                ws.append(self._row(gb))
                total_rows += 1
            for column in ws.columns:
                width = max((len(str(c.value)) for c in column if c.value is not None), default=0)
                ws.column_dimensions[column[0].column_letter].width = min(width + 2, 50)
        if not wb.sheetnames:
            ws = wb.create_sheet('Aucune commune')
            ws.append(REPORT_HEADERS)
        logger.info(
            "account_creation_report user=%s benefit_plan=%s province=%s agency=%s communes=%s rows=%s",
            getattr(self.user, 'login_name', None), benefit_plan_id, province_id,
            payment_agency_id, commune_count, total_rows,
        )
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
