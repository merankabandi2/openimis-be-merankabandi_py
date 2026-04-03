import os
import graphene
from datetime import datetime
from django.conf import settings
from django.core.exceptions import PermissionDenied
from core.schema import OpenIMISMutation
from .models import ResultFrameworkSnapshot, IndicatorAchievement, Indicator, Section
from .result_framework_service import ResultFrameworkService


class CreateResultFrameworkSnapshotInput(graphene.InputObjectType):
    """Input for creating a result framework snapshot"""
    name = graphene.String(required=True)
    description = graphene.String()
    date_from = graphene.Date()
    date_to = graphene.Date()


class UpdateIndicatorAchievementInput(graphene.InputObjectType):
    """Input for updating indicator achievement"""
    indicator_id = graphene.Int(required=True)
    achieved = graphene.Float(required=True)
    date = graphene.Date()
    comment = graphene.String()


class GenerateResultFrameworkDocumentInput(graphene.InputObjectType):
    """Input for generating result framework document"""
    snapshot_id = graphene.ID()
    format = graphene.String(default_value='docx')
    date_from = graphene.Date()
    date_to = graphene.Date()


class CreateResultFrameworkSnapshotMutation(OpenIMISMutation):
    """Create a new result framework snapshot"""
    _mutation_module = "merankabandi"
    _mutation_class = "CreateResultFrameworkSnapshotMutation"

    class Input:
        name = graphene.String(required=True)
        description = graphene.String()
        date_from = graphene.Date()
        date_to = graphene.Date()

    @classmethod
    def async_mutate(cls, user, **data):
        try:
            if not user or user.is_anonymous:
                raise PermissionDenied("User must be authenticated")

            service = ResultFrameworkService()

            snapshot = service.create_snapshot(
                name=data.get('name'),
                description=data.get('description', ''),
                user=user,
                date_from=data.get('date_from'),
                date_to=data.get('date_to')
            )

            return {
                'success': True,
                'message': f'Snapshot created successfully with ID: {snapshot.id}',
                'detail': str(snapshot.id)
            }
        except Exception as e:
            return {
                'success': False,
                'message': str(e),
                'detail': None
            }


class UpdateIndicatorAchievementMutation(OpenIMISMutation):
    """Update or create indicator achievement"""
    _mutation_module = "merankabandi"
    _mutation_class = "UpdateIndicatorAchievementMutation"

    class Input:
        indicator_id = graphene.Int(required=True)
        achieved = graphene.Float(required=True)
        date = graphene.Date()
        comment = graphene.String()

    @classmethod
    def async_mutate(cls, user, **data):
        try:
            if not user or user.is_anonymous:
                raise PermissionDenied("User must be authenticated")

            indicator = Indicator.objects.get(id=data['indicator_id'])

            # Create new achievement record
            achievement = IndicatorAchievement.objects.create(
                indicator=indicator,
                achieved=data['achieved'],
                date=data.get('date') or datetime.now().date(),
                comment=data.get('comment', '')
            )

            return {
                'success': True,
                'message': f'Achievement updated for indicator: {indicator.name}',
                'detail': str(achievement.id)
            }
        except Indicator.DoesNotExist:
            return {
                'success': False,
                'message': 'Indicator not found',
                'detail': None
            }
        except Exception as e:
            return {
                'success': False,
                'message': str(e),
                'detail': None
            }


def _generate_xlsx(sections_data, date_from=None, date_to=None):
    """Generate an xlsx workbook from result framework data."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Cadre de Résultats"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    section_font = Font(name="Calibri", bold=True, size=11, color="2E4057")
    section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    period = ""
    if date_from:
        period += f" du {date_from}"
    if date_to:
        period += f" au {date_to}"
    title_cell.value = f"CADRE DE RÉSULTATS — MERANKABANDI{period}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2E4057")
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"].value = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers row
    headers = ["N°", "Indicateur", "PBC", "Base", "Cible", "Réalisé", "Progression (%)"]
    col_widths = [6, 55, 12, 12, 12, 12, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    row = 5
    indicator_num = 0

    for section in sections_data:
        # Section header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = ws.cell(row=row, column=1, value=section["name"])
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(horizontal="left")
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

        for ind in section["indicators"]:
            indicator_num += 1
            target = float(ind.get("target", 0))
            achieved = float(ind.get("achieved", 0))
            baseline = float(ind.get("baseline", 0))
            progress = (achieved / target * 100) if target > 0 else 0

            values = [
                indicator_num,
                ind.get("name", ""),
                ind.get("pbc", ""),
                baseline,
                target,
                achieved,
                round(progress, 1),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.font = Font(name="Calibri", size=10)
                if col_idx >= 4:
                    cell.alignment = Alignment(horizontal="center")
                    cell.number_format = "#,##0.0" if isinstance(val, float) else "0"

            # Color progress cell
            prog_cell = ws.cell(row=row, column=7)
            if progress >= 80:
                prog_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif progress >= 50:
                prog_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                prog_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            row += 1

    return wb


class GenerateResultFrameworkDocumentMutation(graphene.Mutation):
    """Generate result framework document as xlsx"""
    class Arguments:
        snapshot_id = graphene.ID()
        format = graphene.String(default_value='xlsx')
        date_from = graphene.Date()
        date_to = graphene.Date()

    success = graphene.Boolean()
    message = graphene.String()
    document_url = graphene.String()

    @classmethod
    def mutate(cls, root, info, snapshot_id=None, format='xlsx', date_from=None, date_to=None):
        try:
            user = info.context.user
            if not user or user.is_anonymous:
                raise PermissionDenied("User must be authenticated")

            # Get data — from snapshot or live query
            if snapshot_id:
                snapshot = ResultFrameworkSnapshot.objects.get(id=snapshot_id)
                sections_data = snapshot.data.get("sections", [])
            else:
                service = ResultFrameworkService()
                sections_data = []
                for section in Section.objects.all().prefetch_related("indicators"):
                    section_entry = {"name": section.name, "indicators": []}
                    for indicator in section.indicators.all():
                        result = service.calculate_indicator_value(
                            indicator.id, date_from=date_from, date_to=date_to
                        )
                        section_entry["indicators"].append({
                            "name": indicator.name,
                            "pbc": indicator.pbc or "",
                            "baseline": float(indicator.baseline) if indicator.baseline else 0,
                            "target": float(indicator.target) if indicator.target else 0,
                            "achieved": result.get("value", 0),
                        })
                    sections_data.append(section_entry)

            # Generate xlsx
            wb = _generate_xlsx(sections_data, date_from, date_to)

            # Save to MEDIA_ROOT
            doc_dir = os.path.join(settings.MEDIA_ROOT, "result_framework_docs")
            os.makedirs(doc_dir, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"cadre_resultats_{timestamp}.xlsx"
            filepath = os.path.join(doc_dir, filename)
            wb.save(filepath)

            # Return URL relative to MEDIA_URL
            document_url = f"result_framework_docs/{filename}"

            return cls(
                success=True,
                message="Document generated successfully",
                document_url=document_url,
            )

        except Exception as e:
            return cls(
                success=False,
                message=str(e),
                document_url=None,
            )


class FinalizeSnapshotMutation(OpenIMISMutation):
    """Finalize a snapshot to prevent further changes"""
    _mutation_module = "merankabandi"
    _mutation_class = "FinalizeSnapshotMutation"

    class Input:
        snapshot_id = graphene.ID(required=True)

    @classmethod
    def async_mutate(cls, user, **data):
        try:
            if not user or user.is_anonymous:
                raise PermissionDenied("User must be authenticated")

            snapshot = ResultFrameworkSnapshot.objects.get(id=data['snapshot_id'])

            if snapshot.status != 'DRAFT':
                return {
                    'success': False,
                    'message': 'Snapshot is not in DRAFT status',
                    'detail': None
                }

            snapshot.status = 'FINALIZED'
            snapshot.save()

            return {
                'success': True,
                'message': 'Snapshot finalized successfully',
                'detail': str(snapshot.id)
            }
        except ResultFrameworkSnapshot.DoesNotExist:
            return {
                'success': False,
                'message': 'Snapshot not found',
                'detail': None
            }
        except Exception as e:
            return {
                'success': False,
                'message': str(e),
                'detail': None
            }
