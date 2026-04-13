"""
REST endpoint for sub-component Excel report export.
Generates a multi-sheet XLSX with per-province/commune/colline breakdowns
for each of the 5 program sub-components.
"""
import io
import logging
from datetime import datetime

from django.db import connection
from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated

logger = logging.getLogger(__name__)

# Host community communes (same as models.py)
HOST_COMMUNES = ['Butezi', 'Ruyigi', 'Kiremba', 'Gasorwe', 'Gashoho', 'Muyinga', 'Cankuzo']

# ─── Sheet definitions ────────────────────────────────────────────────────────
SHEETS = [
    {
        'key': 'sc_1_1',
        'title': 'SC 1.1 - Crises El Niño',
        'header': 'Sous-composante 1.1 : Transferts monétaires pour la réponse aux crises éligibles aux ménages affectés par le phénomène El Niño',
        'planned_label': 'Ménages prévus',
        'achieved_label': 'Ménages appuyés',
        'source': 'beneficiary',
        'filters': {"benefit_plan_code": "1.1"},
    },
    {
        'key': 'sc_1_2',
        'title': 'SC 1.2 - TM réguliers',
        'header': 'Sous-composante 1.2 : Transferts monétaires aux bénéficiaires ordinaires',
        'planned_label': 'Bénéficiaires prévus',
        'achieved_label': 'Bénéficiaires payés',
        'source': 'beneficiary',
        'filters': {"benefit_plan_code": "1.2", "exclude_host": True},
    },
    {
        'key': 'sc_1_3',
        'title': 'SC 1.3 - Mesures accomp.',
        'header': "Sous-composante 1.3 : Mesures d'accompagnement pour le changement de comportement pour les investissements dans le capital humain",
        'planned_label': 'Bénéficiaires prévus',
        'achieved_label': 'Bénéficiaires atteints',
        'source': 'activities',
        'filters': {"activity_types": ['SensitizationTraining', 'BehaviorChangePromotion']},
    },
    {
        'key': 'comp_2',
        'title': 'Comp 2 - Inclusion prod.',
        'header': 'Composante 2 : Inclusion productive',
        'planned_label': 'Bénéficiaires prévus',
        'achieved_label': 'Bénéficiaires atteints',
        'source': 'activities',
        'filters': {"activity_types": ['MicroProject']},
    },
    {
        'key': 'comp_4',
        'title': "Comp 4 - Comm. d'accueil",
        'header': "Composante 4 : Intégration des communautés d'accueil dans les systèmes nationaux de protection sociale",
        'planned_label': 'Bénéficiaires prévus',
        'achieved_label': 'Bénéficiaires payés',
        'source': 'beneficiary',
        'filters': {"benefit_plan_code": "1.2", "host_only": True},
    },
]


def _dictfetchall(cursor):
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _query_beneficiary_data(filters):
    """Query dashboard_beneficiary_summary for location-level aggregates."""
    conditions = ['"isDeleted" = false']
    params = []

    if filters.get('benefit_plan_code'):
        conditions.append('benefit_plan_code = %s')
        params.append(filters['benefit_plan_code'])

    if filters.get('exclude_host'):
        placeholders = ','.join(['%s'] * len(HOST_COMMUNES))
        conditions.append(f"(community_type IS NULL OR community_type != 'HOST')")

    if filters.get('host_only'):
        conditions.append("community_type = 'HOST'")

    where = ' AND '.join(conditions)

    query = f"""
        SELECT
            province, commune, colline,
            SUM(beneficiary_count) AS total,
            SUM(male_count) AS male,
            SUM(female_count) AS female,
            SUM(twa_count) AS twa
        FROM dashboard_beneficiary_summary
        WHERE {where}
        GROUP BY province, commune, colline
        ORDER BY province, commune, colline
    """

    with connection.cursor() as cursor:
        cursor.execute(query, params)
        return _dictfetchall(cursor)


def _query_activity_data(filters):
    """Query dashboard_activities_summary for location-level aggregates."""
    activity_types = filters.get('activity_types', [])
    placeholders = ','.join(['%s'] * len(activity_types))

    query = f"""
        SELECT
            province_name AS province,
            commune_name AS commune,
            location_name AS colline,
            SUM(total_participants) AS total,
            SUM(male_participants) AS male,
            SUM(female_participants) AS female,
            SUM(twa_participants) AS twa
        FROM dashboard_activities_summary
        WHERE activity_type IN ({placeholders})
        GROUP BY province_name, commune_name, location_name
        ORDER BY province_name, commune_name, location_name
    """

    with connection.cursor() as cursor:
        cursor.execute(query, activity_types)
        return _dictfetchall(cursor)


def _build_workbook(sheets_data):
    """Build an openpyxl Workbook with one sheet per sub-component."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, size=12)
    subheader_font = Font(bold=True, size=10)
    col_font = Font(bold=True, size=9)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    for idx, sheet_def in enumerate(SHEETS):
        key = sheet_def['key']
        data = sheets_data.get(key, [])
        ws = wb.create_sheet(title=sheet_def['title'])

        # Row 1: empty
        # Row 2: section header — write value BEFORE merging to avoid MergedCell error
        row = 2
        cell = ws.cell(row=row, column=1, value=sheet_def['header'])
        cell.font = header_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=11)

        # Row 4-5: column headers (two-level)
        # Write all cell values and formatting BEFORE merging to avoid
        # openpyxl MergedCell read-only errors.
        row = 4
        planned_label = sheet_def['planned_label']
        achieved_label = sheet_def['achieved_label']

        headers_r1 = ['Province', 'Commune', 'Colline', planned_label, '', 'TOTAL', 'Twa',
                       achieved_label, '', 'TOTAL', 'Twa']
        headers_r2 = ['', '', '', 'Homme', 'Femme', '', '', 'Homme', 'Femme', '', '']

        for col, val in enumerate(headers_r1, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = col_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = header_fill
            c.border = thin_border

        for col, val in enumerate(headers_r2, 1):
            c = ws.cell(row=row + 1, column=col, value=val)
            c.font = col_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.fill = header_fill
            c.border = thin_border

        # Now merge cells after all values are written
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)  # Province
        ws.merge_cells(start_row=row, start_column=2, end_row=row + 1, end_column=2)  # Commune
        ws.merge_cells(start_row=row, start_column=3, end_row=row + 1, end_column=3)  # Colline
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)       # Planned H/F
        ws.merge_cells(start_row=row, start_column=6, end_row=row + 1, end_column=6)  # TOTAL
        ws.merge_cells(start_row=row, start_column=7, end_row=row + 1, end_column=7)  # Twa
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)       # Achieved H/F
        ws.merge_cells(start_row=row, start_column=10, end_row=row + 1, end_column=10) # TOTAL
        ws.merge_cells(start_row=row, start_column=11, end_row=row + 1, end_column=11) # Twa

        # Data rows starting at row 6
        data_row = row + 2
        for record in data:
            male = record.get('male', 0) or 0
            female = record.get('female', 0) or 0
            total = record.get('total', 0) or 0
            twa = record.get('twa', 0) or 0

            values = [
                record.get('province', ''),
                record.get('commune', ''),
                record.get('colline', ''),
                male,           # Planned Homme
                female,         # Planned Femme
                total,          # Planned TOTAL
                twa,            # Planned Twa
                '',             # Achieved Homme (to be filled manually or from payment data)
                '',             # Achieved Femme
                '',             # Achieved TOTAL
                '',             # Achieved Twa
            ]

            for col, val in enumerate(values, 1):
                c = ws.cell(row=data_row, column=col, value=val)
                c.border = thin_border
                if isinstance(val, (int, float)):
                    c.number_format = '#,##0'

            data_row += 1

        # Column widths
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 18
        for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            ws.column_dimensions[col_letter].width = 12

    return wb


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_subcomponents_report(request):
    """
    GET /api/merankabandi/export/subcomponents/
    Returns a multi-sheet Excel file with per-sub-component beneficiary data.
    """
    try:
        sheets_data = {}

        for sheet_def in SHEETS:
            if sheet_def['source'] == 'beneficiary':
                sheets_data[sheet_def['key']] = _query_beneficiary_data(sheet_def['filters'])
            elif sheet_def['source'] == 'activities':
                sheets_data[sheet_def['key']] = _query_activity_data(sheet_def['filters'])

        wb = _build_workbook(sheets_data)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        today = datetime.now().strftime('%Y-%m-%d')
        filename = f'rapport_sous_composantes_{today}.xlsx'

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.document',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        logger.error(f"Subcomponents export error: {e}")
        from rest_framework.response import Response
        from rest_framework import status
        return Response({'success': False, 'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─── Result Framework Export ─────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_result_framework(request):
    """Export the Result Framework as a styled xlsx.

    Uses the latest snapshot if available and less than 1 year old.
    Otherwise computes fresh data (slow). The snapshot date is included
    in the export filename and title.
    """
    from .models import ResultFrameworkSnapshot, Section
    from .result_framework_service import ResultFrameworkService
    from .result_framework_mutations import _generate_xlsx
    from django.utils import timezone
    from dateutil.relativedelta import relativedelta

    date_from = request.query_params.get('date_from')
    date_to = request.query_params.get('date_to')
    force_fresh = request.query_params.get('fresh', '').lower() in ('1', 'true')

    try:
        snapshot = None
        snapshot_date_str = None

        if not force_fresh:
            # Try to use latest snapshot (must be less than 1 month old)
            one_month_ago = timezone.now() - relativedelta(months=1)
            snapshot = (
                ResultFrameworkSnapshot.objects
                .filter(snapshot_date__gte=one_month_ago)
                .order_by('-snapshot_date')
                .first()
            )

        if snapshot and snapshot.data:
            # Use snapshot data (fast) — serve xlsx if already generated
            if snapshot.document_path:
                import os
                filepath = os.path.join(getattr(settings, 'MEDIA_ROOT', 'file_storage'), snapshot.document_path)
                if os.path.exists(filepath):
                    with open(filepath, 'rb') as f:
                        response = HttpResponse(
                            f.read(),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        )
                    response['Content-Disposition'] = f'attachment; filename="{os.path.basename(filepath)}"'
                    return response

            # Snapshot exists but no xlsx — generate from snapshot data
            sections_data = snapshot.data.get('sections', [])
            snapshot_date_str = snapshot.snapshot_date.strftime('%d/%m/%Y')
            logger.info(f"Result framework export using snapshot from {snapshot_date_str}")

            wb = _generate_xlsx(sections_data, date_from, date_to, snapshot_date=snapshot_date_str)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            response = HttpResponse(
                buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="cadre_resultats_{snapshot.snapshot_date.strftime("%Y%m%d")}.xlsx"'
            return response
        else:
            # No recent snapshot — trigger async creation, return 202
            from .tasks import create_result_framework_snapshot
            from rest_framework.response import Response

            user = request.user
            create_result_framework_snapshot.delay(str(user.id))
            logger.info(f"No recent snapshot — triggered async creation for user {user}")

            return Response({
                'status': 'processing',
                'message': 'Aucun snapshot récent. Le cadre de résultats est en cours de génération. '
                           'Vous recevrez une notification quand il sera prêt.',
            }, status=202)

    except Exception as e:
        logger.error(f"Result framework export error: {e}", exc_info=True)
        from rest_framework.response import Response
        from rest_framework import status as rf_status
        return Response({'success': False, 'error': str(e)}, status=rf_status.HTTP_500_INTERNAL_SERVER_ERROR)


# ─── Activity Table Exports (xlsx) ───────────────────────────────────────────

def _activity_xlsx(queryset, columns, sheet_title, filename):
    """Generic xlsx builder for activity tables."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='2E4057', end_color='2E4057', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Headers
    for col_idx, (label, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    for row_idx, obj in enumerate(queryset, 2):
        for col_idx, (_, getter, _) in enumerate(columns, 1):
            val = getter(obj)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_sensitization_trainings(request):
    """Export Sensitization/Formation data as xlsx."""
    from .models import SensitizationTraining

    qs = SensitizationTraining.objects.select_related(
        'location', 'location__parent', 'location__parent__parent',
    ).order_by('-sensitization_date')

    # Apply filters from query params
    province_id = request.query_params.get('province_id')
    if province_id:
        qs = qs.filter(location__parent__parent__id=province_id)
    category = request.query_params.get('category')
    if category:
        qs = qs.filter(category__icontains=category)
    validation_status = request.query_params.get('validation_status')
    if validation_status:
        qs = qs.filter(validation_status=validation_status)
    facilitator = request.query_params.get('facilitator')
    if facilitator:
        qs = qs.filter(facilitator__icontains=facilitator)

    def _get_category_label(obj):
        labels = dict(SensitizationTraining.THEME_CATEGORIES)
        return labels.get(obj.category, obj.category or '')

    def _get_modules(obj):
        if not obj.modules:
            return ''
        return ', '.join(obj.modules)

    columns = [
        ('Date', lambda o: str(o.sensitization_date) if o.sensitization_date else '', 12),
        ('Province', lambda o: o.location.parent.parent.name if o.location and o.location.parent and o.location.parent.parent else '', 15),
        ('Commune', lambda o: o.location.parent.name if o.location and o.location.parent else '', 15),
        ('Colline', lambda o: o.location.name if o.location else '', 15),
        ('Catégorie', _get_category_label, 30),
        ('Thèmes', _get_modules, 30),
        ('Facilitateur', lambda o: o.facilitator or '', 20),
        ('Hommes', lambda o: o.male_participants, 10),
        ('Femmes', lambda o: o.female_participants, 10),
        ('Twa', lambda o: o.twa_participants, 10),
        ('Observations', lambda o: o.observations or '', 25),
        ('Statut', lambda o: o.validation_status or '', 12),
    ]

    timestamp = datetime.now().strftime('%Y%m%d')
    return _activity_xlsx(qs, columns, 'Sensibilisations', f'sensibilisations_{timestamp}.xlsx')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_micro_projects(request):
    """Export Micro-projets data as xlsx."""
    from .models import MicroProject
    from django.db.models import Max

    qs = MicroProject.objects.select_related(
        'location', 'location__parent', 'location__parent__parent',
    ).order_by('location__name', '-report_date')

    province_id = request.query_params.get('province_id')
    if province_id:
        qs = qs.filter(location__parent__parent__id=province_id)
    validation_status = request.query_params.get('validation_status')
    if validation_status:
        qs = qs.filter(validation_status=validation_status)

    # Identify latest report per colline
    latest_dates = {
        r['location_id']: r['latest_date']
        for r in qs.values('location_id').annotate(latest_date=Max('report_date'))
    }

    columns = [
        ('Date du rapport', lambda o: str(o.report_date) if o.report_date else '', 12),
        ('Province', lambda o: o.location.parent.parent.name if o.location and o.location.parent and o.location.parent.parent else '', 15),
        ('Commune', lambda o: o.location.parent.name if o.location and o.location.parent else '', 15),
        ('Colline', lambda o: o.location.name if o.location else '', 15),
        ('Type de projet', lambda o: o.project_type or '', 25),
        ('Hommes', lambda o: o.male_participants, 10),
        ('Femmes', lambda o: o.female_participants, 10),
        ('Twa', lambda o: o.twa_participants, 10),
        ('Dernier', lambda o: 'Oui' if latest_dates.get(o.location_id) == o.report_date else '', 8),
        ('Observations', lambda o: o.observations or '', 25),
        ('Statut', lambda o: o.validation_status or '', 12),
    ]

    timestamp = datetime.now().strftime('%Y%m%d')
    return _activity_xlsx(qs, columns, 'État des micro-projets', f'etat_micro_projets_{timestamp}.xlsx')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_behavior_change_promotions(request):
    """Export Suivi de l'adoption MACH data as xlsx."""
    from .models import BehaviorChangePromotion
    from django.db.models import Max

    qs = BehaviorChangePromotion.objects.select_related(
        'location', 'location__parent', 'location__parent__parent',
    ).order_by('location__name', '-report_date')

    province_id = request.query_params.get('province_id')
    if province_id:
        qs = qs.filter(location__parent__parent__id=province_id)
    validation_status = request.query_params.get('validation_status')
    if validation_status:
        qs = qs.filter(validation_status=validation_status)

    # Identify latest report per colline
    latest_dates = {
        r['location_id']: r['latest_date']
        for r in qs.values('location_id').annotate(latest_date=Max('report_date'))
    }

    columns = [
        ('Date du rapport', lambda o: str(o.report_date) if o.report_date else '', 12),
        ('Province', lambda o: o.location.parent.parent.name if o.location and o.location.parent and o.location.parent.parent else '', 15),
        ('Commune', lambda o: o.location.parent.name if o.location and o.location.parent else '', 15),
        ('Colline', lambda o: o.location.name if o.location else '', 15),
        ('Ménages touchés H', lambda o: o.male_participants, 10),
        ('Ménages touchés F', lambda o: o.female_participants, 10),
        ('Twa', lambda o: o.twa_participants, 10),
        ('Dernier', lambda o: 'Oui' if latest_dates.get(o.location_id) == o.report_date else '', 8),
        ('Observations', lambda o: o.observations or '', 25),
        ('Statut', lambda o: o.validation_status or '', 12),
    ]

    timestamp = datetime.now().strftime('%Y%m%d')
    return _activity_xlsx(qs, columns, "Suivi de l'adoption MACH", f'suivi_adoption_{timestamp}.xlsx')
