import pandas as pd
from datetime import datetime
from django.http import HttpResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import uuid
import logging

from django.db.models import Q
from location.models import Location
from payroll.models import PaymentPoint
from social_protection.models import BenefitPlan
from .models import MonetaryTransfer

logger = logging.getLogger(__name__)


class MonetaryTransferImportService:
    """Service for importing/exporting MonetaryTransfer data from/to Excel"""
    
    # Define column mappings for Excel import
    EXCEL_COLUMNS = {
        'Date des transferts': 'transfer_date',
        'Commune': 'commune',
        'Colline': 'colline',
        'Programme': 'programme',
        'Agence de paiement': 'payment_agency',
        'Femmes prévues': 'planned_women',
        'Hommes prévus': 'planned_men',
        'Twa prévus': 'planned_twa',
        'Femmes payées': 'paid_women',
        'Hommes payés': 'paid_men',
        'Twa payés': 'paid_twa',
        'Montant prévu': 'planned_amount',
        'Montant transféré': 'transferred_amount',
    }
    
    # MIME types for Excel files
    ACCEPTED_MIME_TYPES = {
        'text/csv': lambda f: pd.read_csv(f),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': lambda f: pd.read_excel(f),
        'application/vnd.ms-excel': lambda f: pd.read_excel(f),
    }
    
    @classmethod
    def validate_file(cls, file):
        """Validate file type and extension"""
        # Check MIME type
        if file.content_type not in cls.ACCEPTED_MIME_TYPES:
            return False, "File type not supported. Please upload CSV or Excel file."
        
        # Check extension
        file_ext = file.name.split('.')[-1].lower()
        if file_ext not in ['csv', 'xls', 'xlsx']:
            return False, "File extension not supported. Please upload CSV or Excel file."
        
        return True, None
    
    @classmethod
    def import_from_excel(cls, file, user):
        """Import MonetaryTransfer data from Excel/CSV file"""
        try:
            # Validate file
            valid, error = cls.validate_file(file)
            if not valid:
                return {
                    'success': False,
                    'error': error,
                    'imported': 0,
                    'failed': 0,
                    'invalid_items': []
                }
            
            # Read file
            loader = cls.ACCEPTED_MIME_TYPES[file.content_type]
            df = loader(file)
            
            # Validate columns
            missing_columns = []
            for col in cls.EXCEL_COLUMNS.keys():
                if col not in df.columns:
                    missing_columns.append(col)
            
            if missing_columns:
                return {
                    'success': False,
                    'error': f"Missing required columns: {', '.join(missing_columns)}",
                    'imported': 0,
                    'failed': 0,
                    'invalid_items': []
                }
            
            # Process rows
            imported = 0
            failed = 0
            invalid_items = []
            
            for index, row in df.iterrows():
                try:
                    # Parse date
                    transfer_date = pd.to_datetime(row['Date des transferts']).date()
                    
                    # Find location by commune-colline pair
                    commune_name = str(row['Commune']).strip()
                    colline_name = str(row['Colline']).strip()
                    
                    # Try to find the colline (village) by matching both commune and colline names
                    location = Location.objects.filter(
                        name__iexact=colline_name,
                        parent__name__iexact=commune_name,
                        type='V'  # Village level
                    ).first()
                    
                    if not location:
                        # Try alternative search - sometimes data might have slight variations
                        location = Location.objects.filter(
                            name__icontains=colline_name,
                            parent__name__icontains=commune_name,
                            type='V'
                        ).first()
                    
                    if not location:
                        invalid_items.append({
                            'row': index + 2,  # Excel row number (1-based + header)
                            'error': f"Location not found for Commune: {commune_name}, Colline: {colline_name}",
                            'data': row.to_dict()
                        })
                        failed += 1
                        continue
                    
                    # Find programme (BenefitPlan)
                    programme_name = str(row['Programme']).strip()
                    programme = BenefitPlan.objects.filter(
                        Q(name__iexact=programme_name) | Q(code__iexact=programme_name)
                    ).first()
                    
                    if not programme:
                        invalid_items.append({
                            'row': index + 2,
                            'error': f"Programme not found: {programme_name}",
                            'data': row.to_dict()
                        })
                        failed += 1
                        continue
                    
                    # Find payment agency (PaymentPoint)
                    agency_name = str(row['Agence de paiement']).strip()
                    payment_agency = PaymentPoint.objects.filter(
                        Q(name__iexact=agency_name) | Q(code__iexact=agency_name)
                    ).first()
                    
                    if not payment_agency:
                        invalid_items.append({
                            'row': index + 2,
                            'error': f"Payment agency not found: {agency_name}",
                            'data': row.to_dict()
                        })
                        failed += 1
                        continue
                    
                    # Create MonetaryTransfer object
                    monetary_transfer = MonetaryTransfer(
                        id=uuid.uuid4(),
                        transfer_date=transfer_date,
                        location=location,
                        programme=programme,
                        payment_agency=payment_agency,
                        planned_women=int(row['Femmes prévues'] or 0),
                        planned_men=int(row['Hommes prévus'] or 0),
                        planned_twa=int(row['Twa prévus'] or 0),
                        paid_women=int(row['Femmes payées'] or 0),
                        paid_men=int(row['Hommes payés'] or 0),
                        paid_twa=int(row['Twa payés'] or 0),
                        planned_amount=float(row['Montant prévu'] or 0),
                        transferred_amount=float(row['Montant transféré'] or 0),
                    )
                    
                    # Validate data
                    if monetary_transfer.paid_women > monetary_transfer.planned_women:
                        invalid_items.append({
                            'row': index + 2,
                            'error': "Paid women cannot exceed planned women",
                            'data': row.to_dict()
                        })
                        failed += 1
                        continue
                    
                    if monetary_transfer.paid_men > monetary_transfer.planned_men:
                        invalid_items.append({
                            'row': index + 2,
                            'error': "Paid men cannot exceed planned men",
                            'data': row.to_dict()
                        })
                        failed += 1
                        continue
                    
                    if monetary_transfer.paid_twa > monetary_transfer.planned_twa:
                        invalid_items.append({
                            'row': index + 2,
                            'error': "Paid Twa cannot exceed planned Twa",
                            'data': row.to_dict()
                        })
                        failed += 1
                        continue
                    
                    if monetary_transfer.transferred_amount > monetary_transfer.planned_amount:
                        invalid_items.append({
                            'row': index + 2,
                            'error': "Transferred amount cannot exceed planned amount",
                            'data': row.to_dict()
                        })
                        failed += 1
                        continue
                    
                    # Save the object
                    monetary_transfer.save()
                    imported += 1
                    
                except Exception as e:
                    invalid_items.append({
                        'row': index + 2,
                        'error': str(e),
                        'data': row.to_dict()
                    })
                    failed += 1
            
            return {
                'success': True,
                'imported': imported,
                'failed': failed,
                'invalid_items': invalid_items[:100]  # Limit to first 100 errors
            }
            
        except Exception as e:
            logger.error(f"Error importing MonetaryTransfer data: {str(e)}")
            return {
                'success': False,
                'error': str(e),
                'imported': 0,
                'failed': 0,
                'invalid_items': []
            }
    
    @classmethod
    def export_to_excel(cls, queryset=None, filters=None):
        """Export MonetaryTransfer data to Excel"""
        try:
            # Get queryset if not provided
            if queryset is None:
                queryset = MonetaryTransfer.objects.all()
                
            # Apply filters if provided
            if filters:
                if 'start_date' in filters and filters['start_date']:
                    queryset = queryset.filter(transfer_date__gte=filters['start_date'])
                if 'end_date' in filters and filters['end_date']:
                    queryset = queryset.filter(transfer_date__lte=filters['end_date'])
                if 'location_id' in filters and filters['location_id']:
                    location = Location.objects.get(id=filters['location_id'])
                    if location.type == 'D':  # Province
                        queryset = queryset.filter(location__parent__parent=location)
                    elif location.type == 'W':  # Commune
                        queryset = queryset.filter(location__parent=location)
                    elif location.type == 'V':  # Colline
                        queryset = queryset.filter(location=location)
                if 'programme_id' in filters and filters['programme_id']:
                    queryset = queryset.filter(programme_id=filters['programme_id'])
                if 'payment_agency_id' in filters and filters['payment_agency_id']:
                    queryset = queryset.filter(payment_agency_id=filters['payment_agency_id'])
            
            # Select related to optimize queries
            queryset = queryset.select_related('location', 'location__parent', 'programme', 'payment_agency')
            
            # Prepare data for export
            data = []
            for transfer in queryset:
                data.append({
                    'Date des transferts': transfer.transfer_date.strftime('%Y-%m-%d'),
                    'Commune': transfer.location.parent.name if transfer.location and transfer.location.parent else '',
                    'Colline': transfer.location.name if transfer.location else '',
                    'Programme': transfer.programme.name if transfer.programme else '',
                    'Agence de paiement': transfer.payment_agency.name if transfer.payment_agency else '',
                    'Femmes prévues': transfer.planned_women,
                    'Hommes prévus': transfer.planned_men,
                    'Twa prévus': transfer.planned_twa,
                    'Total prévus': transfer.total_planned,
                    'Femmes payées': transfer.paid_women,
                    'Hommes payés': transfer.paid_men,
                    'Twa payés': transfer.paid_twa,
                    'Total payés': transfer.total_paid,
                    'Montant prévu': float(transfer.planned_amount),
                    'Montant transféré': float(transfer.transferred_amount),
                    'Taux de paiement (%)': round(
                        (transfer.total_paid / transfer.total_planned * 100) 
                        if transfer.total_planned > 0 else 0, 2
                    ),
                    'Taux de transfert (%)': round(
                        (transfer.transferred_amount / transfer.planned_amount * 100) 
                        if transfer.planned_amount > 0 else 0, 2
                    ),
                })
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Create Excel file with styling
            wb = Workbook()
            ws = wb.active
            ws.title = 'Transferts Monétaires'
            
            # Add header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Style header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            excel_file = BytesIO()
            wb.save(excel_file)
            excel_file.seek(0)
            
            # Create response
            response = HttpResponse(
                excel_file.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'transferts_monetaires_{timestamp}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            logger.error(f"Error exporting MonetaryTransfer data: {str(e)}")
            raise
    
    @classmethod
    def get_import_template(cls):
        """Generate Excel template for import"""
        # Create template data
        template_data = {
            'Date des transferts': ['2024-01-15', '2024-01-20'],
            'Commune': ['Butezi', 'Ruyigi'],
            'Colline': ['Colline 1', 'Colline 2'],
            'Programme': ['Programme 1', 'Programme 2'],
            'Agence de paiement': ['Lumicash', 'Interbank'],
            'Femmes prévues': [50, 75],
            'Hommes prévus': [45, 80],
            'Twa prévus': [5, 10],
            'Femmes payées': [48, 72],
            'Hommes payés': [44, 78],
            'Twa payés': [5, 9],
            'Montant prévu': [2500000, 3750000],
            'Montant transféré': [2400000, 3600000],
        }
        
        # Create DataFrame
        df = pd.DataFrame(template_data)
        
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = 'Template'
        
        # Add header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add instructions sheet
        ws2 = wb.create_sheet('Instructions')
        instructions = [
            ['Instructions pour l\'importation des transferts monétaires'],
            [''],
            ['1. Remplissez toutes les colonnes obligatoires'],
            ['2. Date des transferts: Format YYYY-MM-DD (ex: 2024-01-15)'],
            ['3. Commune et Colline: Doivent correspondre exactement aux noms dans le système'],
            ['4. Programme: Nom ou code du programme existant'],
            ['5. Agence de paiement: Nom de l\'agence existant dans le système'],
            ['6. Les nombres payés ne doivent pas dépasser les nombres prévus'],
            [''],
            ['Colonnes obligatoires:'],
            ['- Date des transferts'],
            ['- Commune'],
            ['- Colline'],
            ['- Programme'],
            ['- Agence de paiement'],
            ['- Au moins une valeur prévue (Femmes/Hommes/Twa)'],
        ]
        
        for row in instructions:
            ws2.append(row)
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        response['Content-Disposition'] = 'attachment; filename="template_transferts_monetaires.xlsx"'
        
        return response