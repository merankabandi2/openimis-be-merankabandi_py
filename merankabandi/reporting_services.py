"""
M&E Reporting Services for Merankabandi Project
Provides dashboard data, Excel exports, and automated aggregation capabilities
"""

import pandas as pd
from datetime import datetime, date
from django.db.models import Sum, Count, Q, F, Case, When, Value, CharField
from django.db.models.functions import Extract, Coalesce
from django.http import HttpResponse
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from .models import (
    SensitizationTraining, BehaviorChangePromotion, MicroProject,
    MonetaryTransfer, Section, Indicator, IndicatorAchievement,
    HOST_COMMUNES
)
from social_protection.models import GroupBeneficiary, BeneficiaryStatus
from location.models import Location


class MEDashboardService:
    """Service for M&E Dashboard data aggregation"""

    HOST_COMMUNES = HOST_COMMUNES

    @classmethod
    def get_beneficiary_breakdown_data(cls, start_date=None, end_date=None, location_id=None):
        """Get beneficiary data broken down by gender, community type, location"""
        
        # Base queryset for monetary transfers
        transfers = MonetaryTransfer.objects.all()
        
        if start_date:
            transfers = transfers.filter(transfer_date__gte=start_date)
        if end_date:
            transfers = transfers.filter(transfer_date__lte=end_date)
        if location_id:
            # Filter by province, commune, or colline
            location = Location.objects.get(id=location_id)
            if location.type == 'D':  # Province
                transfers = transfers.filter(location__parent__parent=location)
            elif location.type == 'W':  # Commune
                transfers = transfers.filter(location__parent=location)
            elif location.type == 'V':  # Colline
                transfers = transfers.filter(location=location)
        
        # Aggregate data with community type annotation
        breakdown = transfers.annotate(
            community_type=Case(
                When(location__parent__name__in=cls.HOST_COMMUNES, then=Value('HOST')),
                default=Value('REFUGEE'),
                output_field=CharField()
            )
        ).aggregate(
            total_planned_men=Sum('planned_men'),
            total_planned_women=Sum('planned_women'),
            total_planned_twa=Sum('planned_twa'),
            total_paid_men=Sum('paid_men'),
            total_paid_women=Sum('paid_women'),
            total_paid_twa=Sum('paid_twa'),
        )
        
        # Get breakdown by community type
        community_breakdown = transfers.values('location__parent__name').annotate(
            community_type=Case(
                When(location__parent__name__in=cls.HOST_COMMUNES, then=Value('HOST')),
                default=Value('REFUGEE'),
                output_field=CharField()
            ),
            planned_men=Sum('planned_men'),
            planned_women=Sum('planned_women'),
            planned_twa=Sum('planned_twa'),
            paid_men=Sum('paid_men'),
            paid_women=Sum('paid_women'),
            paid_twa=Sum('paid_twa'),
        ).order_by('community_type', 'location__parent__name')
        
        return {
            'overall_breakdown': breakdown,
            'community_breakdown': list(community_breakdown),
            'gender_summary': {
                'planned': {
                    'men': breakdown['total_planned_men'] or 0,
                    'women': breakdown['total_planned_women'] or 0,
                    'twa': breakdown['total_planned_twa'] or 0,
                },
                'paid': {
                    'men': breakdown['total_paid_men'] or 0,
                    'women': breakdown['total_paid_women'] or 0,
                    'twa': breakdown['total_paid_twa'] or 0,
                }
            }
        }
    
    @classmethod
    def get_refugee_host_breakdown(cls, start_date=None, end_date=None):
        """Get separate breakdown for refugee vs host community data"""
        
        # Base queryset
        transfers = MonetaryTransfer.objects.all()
        
        if start_date:
            transfers = transfers.filter(transfer_date__gte=start_date)
        if end_date:
            transfers = transfers.filter(transfer_date__lte=end_date)
        
        # Host community data
        host_data = transfers.filter(
            location__parent__name__in=cls.HOST_COMMUNES
        ).aggregate(
            planned_beneficiaries=Sum(F('planned_men') + F('planned_women')),
            paid_beneficiaries=Sum(F('paid_men') + F('paid_women')),
            planned_men=Sum('planned_men'),
            planned_women=Sum('planned_women'),
            planned_twa=Sum('planned_twa'),
            paid_men=Sum('paid_men'),
            paid_women=Sum('paid_women'),
            paid_twa=Sum('paid_twa'),
        )
        
        # Refugee community data
        refugee_data = transfers.exclude(
            location__parent__name__in=cls.HOST_COMMUNES
        ).aggregate(
            planned_beneficiaries=Sum(F('planned_men') + F('planned_women')),
            paid_beneficiaries=Sum(F('paid_men') + F('paid_women')),
            planned_men=Sum('planned_men'),
            planned_women=Sum('planned_women'),
            planned_twa=Sum('planned_twa'),
            paid_men=Sum('paid_men'),
            paid_women=Sum('paid_women'),
            paid_twa=Sum('paid_twa'),
        )
        
        return {
            'host_community': host_data,
            'refugee_community': refugee_data,
            'comparison': {
                'total_planned': (host_data['planned_beneficiaries'] or 0) + (refugee_data['planned_beneficiaries'] or 0),
                'total_paid': (host_data['paid_beneficiaries'] or 0) + (refugee_data['paid_beneficiaries'] or 0),
                'host_percentage': round(
                    ((host_data['planned_beneficiaries'] or 0) / 
                     ((host_data['planned_beneficiaries'] or 0) + (refugee_data['planned_beneficiaries'] or 0)) * 100)
                    if (host_data['planned_beneficiaries'] or 0) + (refugee_data['planned_beneficiaries'] or 0) > 0 else 0, 
                    2
                )
            }
        }
    
    @classmethod
    def get_quarterly_rollup_data(cls, year, quarter=None):
        """Aggregate data by quarters/annually"""
        
        # Define quarter date ranges
        quarter_ranges = {
            1: (f'{year}-01-01', f'{year}-03-31'),
            2: (f'{year}-04-01', f'{year}-06-30'), 
            3: (f'{year}-07-01', f'{year}-09-30'),
            4: (f'{year}-10-01', f'{year}-12-31')
        }
        
        if quarter:
            quarters = [quarter]
        else:
            quarters = [1, 2, 3, 4]
        
        quarterly_data = {}
        
        for q in quarters:
            start_date, end_date = quarter_ranges[q]
            
            # Monetary transfers for quarter
            transfers = MonetaryTransfer.objects.filter(
                transfer_date__range=[start_date, end_date]
            ).aggregate(
                total_planned=Sum(F('planned_men') + F('planned_women')),
                total_paid=Sum(F('paid_men') + F('paid_women')),
                transfer_count=Count('id')
            )
            
            # Training activities for quarter
            training = SensitizationTraining.objects.filter(
                sensitization_date__range=[start_date, end_date]
            ).aggregate(
                total_participants=Sum(F('male_participants') + F('female_participants')),
                session_count=Count('id')
            )
            
            # Behavior change activities for quarter
            behavior_change = BehaviorChangePromotion.objects.filter(
                report_date__range=[start_date, end_date]
            ).aggregate(
                total_participants=Sum(F('male_participants') + F('female_participants')),
                activity_count=Count('id')
            )
            
            # Micro projects for quarter
            micro_projects = MicroProject.objects.filter(
                report_date__range=[start_date, end_date]
            ).aggregate(
                total_beneficiaries=Sum(
                    F('agriculture_beneficiaries') + F('livestock_beneficiaries') + 
                    F('commerce_services_beneficiaries')
                ),
                project_count=Count('id')
            )
            
            quarterly_data[f'Q{q}'] = {
                'transfers': transfers,
                'training': training,
                'behavior_change': behavior_change,
                'micro_projects': micro_projects
            }
        
        return quarterly_data
    
    @classmethod
    def get_twa_minority_metrics(cls, start_date=None, end_date=None):
        """Specific metrics for Twa minority tracking"""
        
        # Base filters
        filters = {}
        if start_date:
            filters['transfer_date__gte'] = start_date
        if end_date:
            filters['transfer_date__lte'] = end_date
        
        # Twa participation across all activities
        twa_transfers = MonetaryTransfer.objects.filter(**filters).aggregate(
            planned_twa=Sum('planned_twa'),
            paid_twa=Sum('paid_twa')
        )
        
        training_filters = {}
        if start_date:
            training_filters['sensitization_date__gte'] = start_date
        if end_date:
            training_filters['sensitization_date__lte'] = end_date
            
        twa_training = SensitizationTraining.objects.filter(**training_filters).aggregate(
            twa_participants=Sum('twa_participants')
        )
        # Count sessions with Twa participants separately
        twa_training['sessions_with_twa'] = SensitizationTraining.objects.filter(
            **training_filters, twa_participants__gt=0
        ).count()
        
        behavior_filters = {}
        if start_date:
            behavior_filters['report_date__gte'] = start_date
        if end_date:
            behavior_filters['report_date__lte'] = end_date
            
        twa_behavior = BehaviorChangePromotion.objects.filter(**behavior_filters).aggregate(
            twa_participants=Sum('twa_participants')
        )
        # Count activities with Twa participants separately
        twa_behavior['activities_with_twa'] = BehaviorChangePromotion.objects.filter(
            **behavior_filters, twa_participants__gt=0
        ).count()
        
        project_filters = {}
        if start_date:
            project_filters['report_date__gte'] = start_date
        if end_date:
            project_filters['report_date__lte'] = end_date
            
        twa_projects = MicroProject.objects.filter(**project_filters).aggregate(
            twa_participants=Sum('twa_participants')
        )
        # Count projects with Twa participants separately
        twa_projects['projects_with_twa'] = MicroProject.objects.filter(
            **project_filters, twa_participants__gt=0
        ).count()
        
        # Calculate inclusion rates
        total_transfers = MonetaryTransfer.objects.filter(**filters).aggregate(
            total_planned=Sum(F('planned_men') + F('planned_women'))
        )['total_planned'] or 0
        
        twa_inclusion_rate = 0
        if total_transfers > 0:
            twa_inclusion_rate = round((twa_transfers['planned_twa'] or 0) / total_transfers * 100, 2)
        
        return {
            'transfers': twa_transfers,
            'training': twa_training,
            'behavior_change': twa_behavior,
            'micro_projects': twa_projects,
            'inclusion_rate': twa_inclusion_rate,
            'summary': {
                'total_twa_in_transfers': twa_transfers['planned_twa'] or 0,
                'total_twa_in_training': twa_training['twa_participants'] or 0,
                'total_twa_in_behavior_change': twa_behavior['twa_participants'] or 0,
                'total_twa_in_projects': twa_projects['twa_participants'] or 0,
            }
        }


class ExcelExportService:
    """Service for Excel report generation matching current templates"""
    
    @classmethod
    def export_monetary_transfers_excel(cls, start_date=None, end_date=None, location_id=None):
        """Export monetary transfer data in Excel format matching TRANSFERTS MONETAIRES.xlsx"""
        
        # Get data
        transfers = MonetaryTransfer.objects.select_related('location', 'programme', 'payment_agency')
        
        if start_date:
            transfers = transfers.filter(transfer_date__gte=start_date)
        if end_date:
            transfers = transfers.filter(transfer_date__lte=end_date)
        if location_id:
            location = Location.objects.get(id=location_id)
            if location.type == 'D':  # Province
                transfers = transfers.filter(location__parent__parent=location)
            elif location.type == 'W':  # Commune
                transfers = transfers.filter(location__parent=location)
        
        # Prepare data for Excel
        data = []
        for transfer in transfers:
            quarter = f"T{((transfer.transfer_date.month-1)//3)+1}"
            
            # Safely get location name
            location_name = ''
            if transfer.location and transfer.location.parent:
                location_name = transfer.location.parent.name
            elif transfer.location:
                location_name = transfer.location.name
            
            # Determine transfer type based on programme and location
            transfer_type = "Transferts monétaires ordinaires"
            if location_name and location_name not in MEDashboardService.HOST_COMMUNES:
                transfer_type = "Transferts monétaires aux ménages Réfugiés"
            
            data.append({
                'Anne': transfer.transfer_date.year,
                'Periode': quarter,
                'Localisation': location_name,  # Commune
                'Types des tranferts': transfer_type,
                'Programme': transfer.programme.name if transfer.programme else '',
                'Agence de paiement': transfer.payment_agency.name if transfer.payment_agency else '',
                'Bénéficiaires prévus': transfer.total_planned,
                'Homme prévu': transfer.planned_men,
                'Femme prévue': transfer.planned_women,
                'Twa prévu': transfer.planned_twa,
                'Bénéficiaires payés': transfer.total_paid,
                'Homme payé': transfer.paid_men,
                'Femme payée': transfer.paid_women,
                'Twa payé': transfer.paid_twa,
                'Taux de paiement (%)': round(
                    (transfer.total_paid / transfer.total_planned * 100) 
                    if transfer.total_planned > 0 else 0, 2
                ),
            })
        
        # Create Excel file
        df = pd.DataFrame(data)
        return cls._create_excel_file(df, 'Transferts Monétaires', 'transferts_monetaires')
    
    @classmethod
    def export_accompanying_measures_excel(cls, start_date=None, end_date=None, location_id=None):
        """Export training/sensitization data in Excel format matching MESURES D ACCOMPAGNEMENTS.xlsx"""
        
        # Get training data
        training_data = []
        
        # Sensitization Training
        training = SensitizationTraining.objects.select_related('location')
        if start_date:
            training = training.filter(sensitization_date__gte=start_date)
        if end_date:
            training = training.filter(sensitization_date__lte=end_date)
        if location_id:
            location = Location.objects.get(id=location_id)
            if location.type == 'D':  # Province
                training = training.filter(location__parent__parent=location)
            elif location.type == 'W':  # Commune
                training = training.filter(location__parent=location)
        
        for item in training:
            quarter = f"T{((item.sensitization_date.month-1)//3)+1}"
            
            # Safely get location name
            location_name = ''
            if item.location and item.location.parent:
                location_name = item.location.parent.name
            elif item.location:
                location_name = item.location.name
            
            community_type = "Communautés d'accueil"
            if location_name and location_name in MEDashboardService.HOST_COMMUNES:
                community_type = "Communautés d'accueil"
            else:
                community_type = "Réfugiés"
            
            training_data.append({
                'Anne': item.sensitization_date.year,
                'Periode': quarter,
                'Localisation': location_name,
                'Types des transferts': f"Formation/Sensibilisation - {community_type}",
                'Activité': item.get_category_display() if item.category else 'Formation générale',
                'Animateur': item.facilitator or '',
                'Homme': item.male_participants,
                'Femme': item.female_participants,
                'Twa': item.twa_participants,
                'Total': item.total_participants,
                'Observations': item.observations or '',
            })
        
        # Behavior Change Promotion
        behavior_change = BehaviorChangePromotion.objects.select_related('location')
        if start_date:
            behavior_change = behavior_change.filter(report_date__gte=start_date)
        if end_date:
            behavior_change = behavior_change.filter(report_date__lte=end_date)
        if location_id:
            location = Location.objects.get(id=location_id)
            if location.type == 'D':  # Province
                behavior_change = behavior_change.filter(location__parent__parent=location)
            elif location.type == 'W':  # Commune
                behavior_change = behavior_change.filter(location__parent=location)
        
        for item in behavior_change:
            quarter = f"T{((item.report_date.month-1)//3)+1}"
            
            # Safely get location name
            location_name = ''
            if item.location and item.location.parent:
                location_name = item.location.parent.name
            elif item.location:
                location_name = item.location.name
            
            community_type = "Communautés d'accueil"
            if location_name and location_name in MEDashboardService.HOST_COMMUNES:
                community_type = "Communautés d'accueil"
            else:
                community_type = "Réfugiés"
            
            training_data.append({
                'Anne': item.report_date.year,
                'Periode': quarter,
                'Localisation': location_name,
                'Types des transferts': f"Promotion changement de comportement - {community_type}",
                'Activité': 'Promotion du changement de comportement',
                'Animateur': '',
                'Homme': item.male_participants,
                'Femme': item.female_participants,
                'Twa': item.twa_participants,
                'Total': item.total_beneficiaries,
                'Observations': item.comments or '',
            })
        
        # Create Excel file
        df = pd.DataFrame(training_data)
        return cls._create_excel_file(df, 'Mesures d\'Accompagnement', 'mesures_accompagnement')
    
    @classmethod
    def export_microprojects_excel(cls, start_date=None, end_date=None, location_id=None):
        """Export micro-project data in Excel format matching MICROPROJET.xlsx"""
        
        # Get micro-project data
        projects = MicroProject.objects.select_related('location')
        
        if start_date:
            projects = projects.filter(report_date__gte=start_date)
        if end_date:
            projects = projects.filter(report_date__lte=end_date)
        if location_id:
            location = Location.objects.get(id=location_id)
            if location.type == 'D':  # Province
                projects = projects.filter(location__parent__parent=location)
            elif location.type == 'W':  # Commune
                projects = projects.filter(location__parent=location)
        
        data = []
        for project in projects:
            quarter = f"T{((project.report_date.month-1)//3)+1}"
            
            # Add row for each project type with beneficiaries
            project_types = [
                ('Agriculture', project.agriculture_beneficiaries),
                ('Élevage général', project.livestock_beneficiaries),
                ('Chèvres', project.livestock_goat_beneficiaries),
                ('Porcins', project.livestock_pig_beneficiaries),
                ('Lapins', project.livestock_rabbit_beneficiaries),
                ('Volailles', project.livestock_poultry_beneficiaries),
                ('Bovins', project.livestock_cattle_beneficiaries),
                ('Commerce et services', project.commerce_services_beneficiaries),
            ]
            
            for project_type, beneficiary_count in project_types:
                if beneficiary_count > 0:
                    # Safely get location name
                    location_name = ''
                    if project.location and project.location.parent:
                        location_name = project.location.parent.name
                    elif project.location:
                        location_name = project.location.name
                    
                    data.append({
                        'Anne': project.report_date.year,
                        'Periode': quarter,
                        'Localisation': location_name,
                        'Type de micro-projets appuyés': project_type,
                        'Nombre de bénéficiaires': beneficiary_count,
                        'Participants Hommes': project.male_participants,
                        'Participants Femmes': project.female_participants,
                        'Participants Twa': project.twa_participants,
                        'Total participants': (project.male_participants + 
                                             project.female_participants + 
                                             project.twa_participants),
                    })
            
            # Add other project types
            for other_type in project.other_project_types.all():
                # Safely get location name
                location_name = ''
                if project.location and project.location.parent:
                    location_name = project.location.parent.name
                elif project.location:
                    location_name = project.location.name
                
                data.append({
                    'Anne': project.report_date.year,
                    'Periode': quarter,
                    'Localisation': location_name,
                    'Type de micro-projets appuyés': other_type.name,
                    'Nombre de bénéficiaires': other_type.beneficiary_count,
                    'Participants Hommes': project.male_participants,
                    'Participants Femmes': project.female_participants,
                    'Participants Twa': project.twa_participants,
                    'Total participants': (project.male_participants + 
                                         project.female_participants + 
                                         project.twa_participants),
                })
        
        # Create Excel file
        df = pd.DataFrame(data)
        return cls._create_excel_file(df, 'Micro-Projets', 'microprojects')
    
    @classmethod
    def export_subcomponents_excel(cls, start_date=None, end_date=None, location_id=None):
        """Export the 5 sub-component report — one sheet per sub-component.

        Columns per sheet:
        Province | Commune | Colline | Planned Men | Planned Women | TOTAL | Twa |
        Actual Men | Actual Women | TOTAL | Twa
        """
        from openpyxl.styles import Border, Side

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Common location filter
        loc_filter = Q()
        if location_id:
            try:
                location = Location.objects.get(id=location_id)
                if location.type == 'D':
                    loc_filter = Q(location__parent__parent=location)
                elif location.type == 'W':
                    loc_filter = Q(location__parent=location)
                elif location.type == 'V':
                    loc_filter = Q(location=location)
            except Location.DoesNotExist:
                pass

        date_filter = Q()
        if start_date:
            date_filter &= Q(transfer_date__gte=start_date)
        if end_date:
            date_filter &= Q(transfer_date__lte=end_date)

        activity_date_filter = Q()
        if start_date:
            activity_date_filter &= Q(report_date__gte=start_date)
        if end_date:
            activity_date_filter &= Q(report_date__lte=end_date)

        def _get_location_hierarchy(loc):
            """Return (province, commune, colline) from a colline-level location."""
            colline = loc.name if loc else ''
            commune = loc.parent.name if loc and loc.parent else ''
            province = loc.parent.parent.name if loc and loc.parent and loc.parent.parent else ''
            return province, commune, colline

        def _style_sheet(ws, title_text, actual_label):
            """Add title row and styled headers matching the Excel template."""
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'),
            )
            header_font = Font(bold=True, size=10)
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            center = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Title row
            ws.merge_cells('A1:K1')
            ws['A1'] = title_text
            ws['A1'].font = Font(bold=True, underline='single', size=11)

            # Header row 1 (merged cells)
            row = 3
            for col, val in [(1, 'Province'), (2, 'Commune'), (3, 'Colline')]:
                ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center
                cell.border = thin_border

            # "Bénéficiaires prévus" group
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=4, value='Bénéficiaires prévus')
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

            for col, val in [(6, 'TOTAL'), (7, 'Twa')]:
                ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

            # Actual group
            ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)
            cell = ws.cell(row=row, column=8, value=actual_label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

            for col, val in [(10, 'TOTAL'), (11, 'Twa')]:
                ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

            # Header row 2 (sub-headers)
            row2 = 4
            for col, val in [(4, 'Homme'), (5, 'Femme'), (8, 'Homme'), (9, 'Femme')]:
                cell = ws.cell(row=row2, column=col, value=val)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center

            # Apply borders to all header cells
            for r in range(3, 5):
                for c in range(1, 12):
                    ws.cell(row=r, column=c).border = thin_border

            # Column widths
            for col, width in [(1, 15), (2, 15), (3, 18), (4, 12), (5, 12),
                               (6, 10), (7, 8), (8, 12), (9, 12), (10, 10), (11, 8)]:
                ws.column_dimensions[chr(64 + col)].width = width

            return 5  # data start row

        def _add_transfer_data(ws, transfers, data_start_row):
            """Add MonetaryTransfer data rows aggregated by colline."""
            from django.db.models import Sum as DjSum
            agg = transfers.values(
                'location__parent__parent__name',  # Province
                'location__parent__name',           # Commune
                'location__name',                   # Colline
            ).annotate(
                p_men=Coalesce(DjSum('planned_men'), 0),
                p_women=Coalesce(DjSum('planned_women'), 0),
                p_twa=Coalesce(DjSum('planned_twa'), 0),
                a_men=Coalesce(DjSum('paid_men'), 0),
                a_women=Coalesce(DjSum('paid_women'), 0),
                a_twa=Coalesce(DjSum('paid_twa'), 0),
            ).order_by('location__parent__parent__name', 'location__parent__name', 'location__name')

            row = data_start_row
            for rec in agg:
                ws.cell(row=row, column=1, value=rec['location__parent__parent__name'] or '')
                ws.cell(row=row, column=2, value=rec['location__parent__name'] or '')
                ws.cell(row=row, column=3, value=rec['location__name'] or '')
                ws.cell(row=row, column=4, value=rec['p_men'])
                ws.cell(row=row, column=5, value=rec['p_women'])
                ws.cell(row=row, column=6, value=rec['p_men'] + rec['p_women'])
                ws.cell(row=row, column=7, value=rec['p_twa'])
                ws.cell(row=row, column=8, value=rec['a_men'])
                ws.cell(row=row, column=9, value=rec['a_women'])
                ws.cell(row=row, column=10, value=rec['a_men'] + rec['a_women'])
                ws.cell(row=row, column=11, value=rec['a_twa'])
                row += 1
            return row

        def _add_activity_data(ws, queryset, data_start_row):
            """Add BehaviorChangePromotion or MicroProject data (no planned fields)."""
            from django.db.models import Sum as DjSum
            agg = queryset.values(
                'location__parent__parent__name',
                'location__parent__name',
                'location__name',
            ).annotate(
                a_men=Coalesce(DjSum('male_participants'), 0),
                a_women=Coalesce(DjSum('female_participants'), 0),
                a_twa=Coalesce(DjSum('twa_participants'), 0),
            ).order_by('location__parent__parent__name', 'location__parent__name', 'location__name')

            row = data_start_row
            for rec in agg:
                ws.cell(row=row, column=1, value=rec['location__parent__parent__name'] or '')
                ws.cell(row=row, column=2, value=rec['location__parent__name'] or '')
                ws.cell(row=row, column=3, value=rec['location__name'] or '')
                # Planned columns left empty (no planned data in these models)
                ws.cell(row=row, column=8, value=rec['a_men'])
                ws.cell(row=row, column=9, value=rec['a_women'])
                ws.cell(row=row, column=10, value=rec['a_men'] + rec['a_women'])
                ws.cell(row=row, column=11, value=rec['a_twa'])
                row += 1
            return row

        # ── Sheet 1: Sous-composante 1.1 (El Nino crisis transfers) ──
        ws1 = wb.create_sheet('1.1 Transferts El Nino')
        data_row = _style_sheet(
            ws1,
            '1. Sous-composante 1.1 : Transferts monétaires pour la réponse aux crises '
            'éligibles aux ménages affectés par le phénomène El Nino',
            'Ménages appuyés',
        )
        transfers_11 = MonetaryTransfer.objects.filter(
            loc_filter, date_filter, programme__code='1.1',
        ).select_related('location__parent__parent')
        _add_transfer_data(ws1, transfers_11, data_row)

        # ── Sheet 2: Sous-composante 1.2 (ordinary transfers) ──
        ws2 = wb.create_sheet('1.2 Transferts ordinaires')
        data_row = _style_sheet(
            ws2,
            '2. Sous-composante 1.2 : Transferts monétaires aux bénéficiaires ordinaires',
            'Bénéficiaires payés',
        )
        transfers_12 = MonetaryTransfer.objects.filter(
            loc_filter, date_filter, programme__code='1.2',
        ).select_related('location__parent__parent')
        _add_transfer_data(ws2, transfers_12, data_row)

        # ── Sheet 3: Sous-composante 1.3 (behavior change) ──
        ws3 = wb.create_sheet('1.3 Changement comportement')
        data_row = _style_sheet(
            ws3,
            '3. Sous-composante 1.3 : Mesures d\'accompagnement pour le changement de '
            'comportement pour les investissements dans le capital humain',
            'Bénéficiaires atteints',
        )
        bcp = BehaviorChangePromotion.objects.filter(
            loc_filter, activity_date_filter,
        ).select_related('location__parent__parent')
        _add_activity_data(ws3, bcp, data_row)

        # ── Sheet 4: Composante 2 (productive inclusion / micro-projects) ──
        ws4 = wb.create_sheet('2 Inclusion productive')
        data_row = _style_sheet(
            ws4,
            '4. Composante 2 : Inclusion productive',
            'Bénéficiaires atteints',
        )
        mp = MicroProject.objects.filter(
            loc_filter, activity_date_filter,
        ).select_related('location__parent__parent')
        _add_activity_data(ws4, mp, data_row)

        # ── Sheet 5: Composante 4 (host communities) ──
        ws5 = wb.create_sheet('4 Communautés accueil')
        data_row = _style_sheet(
            ws5,
            '5. Composante 4 : Intégration des communautés d\'accueil dans les '
            'systèmes nationaux de protection sociale',
            'Bénéficiaires payés',
        )
        transfers_host = MonetaryTransfer.objects.filter(
            loc_filter, date_filter,
            location__parent__name__in=HOST_COMMUNES,
        ).select_related('location__parent__parent')
        _add_transfer_data(ws5, transfers_host, data_row)

        # ── Save and return response ──
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'rapport_sous_composantes_{timestamp}.xlsx'
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @classmethod
    def _create_excel_file(cls, df, sheet_name, filename_prefix):
        """Create formatted Excel file from DataFrame"""
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Add header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # Create response
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{filename_prefix}_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


class IndicatorAggregationService:
    """Service for automated indicator achievement aggregation"""
    
    @classmethod
    def auto_aggregate_household_registration(cls):
        """Auto-aggregate household registration numbers from GroupBeneficiary"""
        
        # Get total households registered in social registry
        total_households = GroupBeneficiary.objects.filter(
            status=BeneficiaryStatus.ACTIVE
        ).count()
        
        # Separate by community type
        host_households = GroupBeneficiary.objects.filter(
            status=BeneficiaryStatus.ACTIVE,
            group__location__parent__name__in=MEDashboardService.HOST_COMMUNES
        ).count()
        
        refugee_households = total_households - host_households
        
        # Update indicators
        try:
            # Total households indicator - try to find exact match first
            try:
                total_indicator = Indicator.objects.get(
                    name="Ménages des zones ciblées inscrits au Registre social national"
                )
            except Indicator.DoesNotExist:
                # If exact match not found, try with icontains but get the first one
                total_indicators = Indicator.objects.filter(
                    name__icontains="Ménages des zones ciblées inscrits au Registre social national"
                )
                if not total_indicators.exists():
                    raise Indicator.DoesNotExist("Total households indicator not found")
                total_indicator = total_indicators.first()
            
            IndicatorAchievement.objects.create(
                indicator=total_indicator,
                achieved=total_households,
                date=date.today(),
                comment="Auto-aggregated from GroupBeneficiary records"
            )
            
            # Host community households indicator
            try:
                host_indicator = Indicator.objects.get(
                    name="Ménages des zones ciblées inclus dans le registre social national - communautés d'accueil"
                )
            except Indicator.DoesNotExist:
                host_indicators = Indicator.objects.filter(
                    name__icontains="communautés d'accueil"
                ).filter(
                    name__icontains="Ménages"
                ).filter(
                    name__icontains="registre social"
                )
                if not host_indicators.exists():
                    raise Indicator.DoesNotExist("Host community households indicator not found")
                host_indicator = host_indicators.first()
            
            IndicatorAchievement.objects.create(
                indicator=host_indicator,
                achieved=host_households,
                date=date.today(),
                comment="Auto-aggregated from GroupBeneficiary records (host communities)"
            )
            
            # Refugee households indicator
            try:
                refugee_indicator = Indicator.objects.get(
                    name="Ménages des zones ciblées inclus dans le registre social national - réfugiés"
                )
            except Indicator.DoesNotExist:
                refugee_indicators = Indicator.objects.filter(
                    name__icontains="réfugiés"
                ).filter(
                    name__icontains="Ménages"
                ).filter(
                    name__icontains="registre social"
                )
                if not refugee_indicators.exists():
                    raise Indicator.DoesNotExist("Refugee households indicator not found")
                refugee_indicator = refugee_indicators.first()
            
            IndicatorAchievement.objects.create(
                indicator=refugee_indicator,
                achieved=refugee_households,
                date=date.today(),
                comment="Auto-aggregated from GroupBeneficiary records (refugees)"
            )
            
            return True, "Household registration indicators updated successfully"
            
        except Indicator.DoesNotExist as e:
            return False, f"Required indicator not found: {e}"
        except Exception as e:
            return False, f"Error updating indicators: {e}"
    
    @classmethod
    def auto_aggregate_transfer_beneficiaries(cls, year=None, quarter=None):
        """Auto-aggregate transfer beneficiary numbers"""
        
        if not year:
            year = date.today().year
        
        # Define date range
        if quarter:
            quarter_ranges = {
                1: (f'{year}-01-01', f'{year}-03-31'),
                2: (f'{year}-04-01', f'{year}-06-30'), 
                3: (f'{year}-07-01', f'{year}-09-30'),
                4: (f'{year}-10-01', f'{year}-12-31')
            }
            start_date, end_date = quarter_ranges[quarter]
            filters = {'transfer_date__range': [start_date, end_date]}
        else:
            filters = {'transfer_date__year': year}
        
        # Get transfer data
        transfers = MonetaryTransfer.objects.filter(**filters)
        
        # Calculate aggregates
        total_beneficiaries = transfers.aggregate(
            total=Sum(F('paid_men') + F('paid_women'))
        )['total'] or 0
        
        female_beneficiaries = transfers.aggregate(
            total=Sum('paid_women')
        )['total'] or 0
        
        refugee_beneficiaries = transfers.exclude(
            location__parent__name__in=MEDashboardService.HOST_COMMUNES
        ).aggregate(
            total=Sum(F('paid_men') + F('paid_women'))
        )['total'] or 0
        
        # Update indicators
        try:
            # Total beneficiaries
            try:
                total_indicator = Indicator.objects.get(
                    name="Bénéficiaires des programmes de protection sociale"
                )
            except Indicator.DoesNotExist:
                total_indicators = Indicator.objects.filter(
                    name__icontains="Bénéficiaires des programmes de protection sociale"
                ).exclude(
                    name__icontains="Femmes"
                ).exclude(
                    name__icontains="réfugiés"
                )
                if not total_indicators.exists():
                    raise Indicator.DoesNotExist("Total beneficiaries indicator not found")
                total_indicator = total_indicators.first()
            
            period_desc = f"Q{quarter} {year}" if quarter else str(year)
            IndicatorAchievement.objects.create(
                indicator=total_indicator,
                achieved=total_beneficiaries,
                date=date.today(),
                comment=f"Auto-aggregated from MonetaryTransfer records for {period_desc}"
            )
            
            # Female beneficiaries
            try:
                female_indicator = Indicator.objects.get(
                    name="Bénéficiaires des programmes de protection sociale - Femmes"
                )
            except Indicator.DoesNotExist:
                female_indicators = Indicator.objects.filter(
                    name__icontains="Bénéficiaires"
                ).filter(
                    name__icontains="Femmes"
                ).filter(
                    name__icontains="protection sociale"
                )
                if not female_indicators.exists():
                    raise Indicator.DoesNotExist("Female beneficiaries indicator not found")
                female_indicator = female_indicators.first()
            
            IndicatorAchievement.objects.create(
                indicator=female_indicator,
                achieved=female_beneficiaries,
                date=date.today(),
                comment=f"Auto-aggregated from MonetaryTransfer records for {period_desc}"
            )
            
            # Refugee beneficiaries
            try:
                refugee_indicator = Indicator.objects.get(
                    name="Bénéficiaires des programmes de filets de sécurité - réfugiés"
                )
            except Indicator.DoesNotExist:
                refugee_indicators = Indicator.objects.filter(
                    name__icontains="Bénéficiaires"
                ).filter(
                    name__icontains="réfugiés"
                ).filter(
                    Q(name__icontains="filets de sécurité") | Q(name__icontains="protection sociale")
                )
                if not refugee_indicators.exists():
                    raise Indicator.DoesNotExist("Refugee beneficiaries indicator not found")
                refugee_indicator = refugee_indicators.first()
            
            IndicatorAchievement.objects.create(
                indicator=refugee_indicator,
                achieved=refugee_beneficiaries,
                date=date.today(),
                comment=f"Auto-aggregated from MonetaryTransfer records for {period_desc}"
            )
            
            return True, f"Transfer beneficiary indicators updated for {period_desc}"
            
        except Indicator.DoesNotExist as e:
            return False, f"Required indicator not found: {e}"
        except Exception as e:
            return False, f"Error updating indicators: {e}"
    
    @classmethod
    def calculate_achievement_by_dimensions(cls, indicator_id, **dimensions):
        """Calculate achievements by location, gender, community type"""
        
        indicator = Indicator.objects.get(id=indicator_id)
        
        # Base queryset depending on indicator type
        if "transfert" in indicator.name.lower() or "bénéficiaire" in indicator.name.lower():
            queryset = MonetaryTransfer.objects.all()
            value_field = F('paid_men') + F('paid_women')
        elif "formation" in indicator.name.lower() or "sensibilisation" in indicator.name.lower():
            queryset = SensitizationTraining.objects.all()
            value_field = F('male_participants') + F('female_participants') + F('twa_participants')
        else:
            return None
        
        # Apply dimension filters
        if 'location_id' in dimensions:
            location = Location.objects.get(id=dimensions['location_id'])
            if location.type == 'D':  # Province
                queryset = queryset.filter(location__parent__parent=location)
            elif location.type == 'W':  # Commune
                queryset = queryset.filter(location__parent=location)
            elif location.type == 'V':  # Colline
                queryset = queryset.filter(location=location)
        
        if 'community_type' in dimensions:
            if dimensions['community_type'] == 'HOST':
                queryset = queryset.filter(location__parent__name__in=MEDashboardService.HOST_COMMUNES)
            elif dimensions['community_type'] == 'REFUGEE':
                queryset = queryset.exclude(location__parent__name__in=MEDashboardService.HOST_COMMUNES)
        
        if 'start_date' in dimensions:
            date_field = 'transfer_date' if hasattr(queryset.model, 'transfer_date') else 'sensitization_date'
            if hasattr(queryset.model, 'report_date'):
                date_field = 'report_date'
            queryset = queryset.filter(**{f'{date_field}__gte': dimensions['start_date']})
        
        if 'end_date' in dimensions:
            date_field = 'transfer_date' if hasattr(queryset.model, 'transfer_date') else 'sensitization_date'
            if hasattr(queryset.model, 'report_date'):
                date_field = 'report_date'
            queryset = queryset.filter(**{f'{date_field}__lte': dimensions['end_date']})
        
        # Calculate total
        result = queryset.aggregate(total=Sum(value_field))
        
        return result['total'] or 0